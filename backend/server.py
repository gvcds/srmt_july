from fastapi import FastAPI, HTTPException, Depends, UploadFile, File
from pydantic import BaseModel
from typing import Optional, Dict, Any, List, Tuple
import re
from sqlalchemy import create_engine, Column, Integer, String, UniqueConstraint, Text, DateTime, ForeignKey, CheckConstraint, or_, inspect, text
from sqlalchemy.orm import sessionmaker, declarative_base, Session
from sqlalchemy.engine import URL
from fastapi.middleware.cors import CORSMiddleware
from datetime import datetime, date
import json
import os
import requests
import xml.etree.ElementTree as ET
import html
import concurrent.futures
import pandas as pd
import io
import base64
import psycopg2
from psycopg2.extras import DictCursor
from psycopg2 import pool
import urllib3
import time
import difflib
from openpyxl.cell.rich_text import TextBlock, CellRichText
from openpyxl.cell.text import InlineFont

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

try:
    from ldap3 import Server, Connection
except Exception:  # pragma: no cover - ambiente sem ldap3
    Server = None
    Connection = None

# --- CONFIGURAÇÃO DO BANCO DE DADOS (SRMT) ---
DB_HOST = "localhost"
DB_PORT = "5432"
DB_NAME = "srmt"
DB_USER = "postgres"
DB_PASS = "326741@Vv"

# --- CONFIGURAÇÃO LDAP ---
LDAP_DOMAIN = "corp"
LDAP_DC_HOST = "105.203.200.250"
LDAP_DNS = "HQBRDC001"

url_object = URL.create(
    "postgresql+psycopg2",
    username=DB_USER,
    password=DB_PASS,
    host=DB_HOST,
    port=int(DB_PORT),
    database=DB_NAME,
)

engine = create_engine(url_object)
SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
Base = declarative_base()

# --- MODELOS ---

class User(Base):
    __tablename__ = "users"
    id = Column(Integer, primary_key=True, index=True)
    first_name = Column(String, nullable=False)
    last_name = Column(String, nullable=False)
    email = Column(String, unique=True, index=True, nullable=False)
    password = Column(String, nullable=False)
    role = Column(String)
    department = Column(String)
    team = Column(String)
    cell = Column(String, nullable=True)
    kp = Column(String, nullable=True) 
    kp_type = Column(String, nullable=True) # 'projeto' ou 'especialista'
    is_backup = Column(Integer, default=0) # 0 para não, 1 para sim
    is_specialist = Column(Integer, default=0) # 0 para não, 1 para sim
    avatar = Column(Text, nullable=True)
    bio = Column(Text, nullable=True)
    skills = Column(Text, nullable=True)
    created_at = Column(DateTime, default=datetime.utcnow)

class Suggestion(Base):
    __tablename__ = "suggestions"
    id = Column(Integer, primary_key=True, index=True)
    field_type = Column(String, index=True)
    value = Column(String, index=True)
    __table_args__ = (UniqueConstraint('field_type', 'value', name='_field_value_uc'),)

class Report(Base):
    __tablename__ = "reports"
    id = Column(Integer, primary_key=True, index=True)
    team = Column(String)
    tester_id = Column(String)
    full_data = Column(Text)
    created_at = Column(DateTime, default=datetime.utcnow)

class Vacation(Base):
    __tablename__ = "vacations"
    id = Column(Integer, primary_key=True, index=True)
    user_id = Column(Integer, ForeignKey("users.id"), index=True, nullable=False)
    start_date = Column(String, nullable=False)
    end_date = Column(String, nullable=False)
    category = Column(String, default="vacation") # 'vacation', 'urgent', 'day-off'
    status = Column(String, default="pending", index=True)
    sell_days = Column(Integer, default=0) # 0 para não, 1 para sim (venda de 10 dias)
    created_at = Column(DateTime, default=datetime.utcnow)
    updated_at = Column(DateTime, default=datetime.utcnow)
    __table_args__ = (
        UniqueConstraint("user_id", "start_date", "end_date", name="uq_vacation_user_period"),
        CheckConstraint("status in ('pending', 'approved', 'rejected', 'conflict', 'fluig_approved', 'downloaded')", name="ck_vacation_status"),
    )

class Ticket(Base):
    __tablename__ = "tickets"
    id = Column(Integer, primary_key=True, index=True)
    type = Column(String, index=True) # 'ticket', 'improvement', 'project'
    priority = Column(String, default="Média") # 'Baixa', 'Média', 'Alta', 'Urgente'
    status = Column(String, default="pendente") # 'pendente', 'aceito', 'rejeitado', 'concluido'
    title = Column(String)
    content = Column(Text)
    resolution = Column(Text, nullable=True) # Texto de justificativa da conclusão
    creators = Column(Text) # JSON list of creators
    created_at = Column(DateTime, default=datetime.utcnow)

class Notification(Base):
    __tablename__ = "notifications"
    id = Column(Integer, primary_key=True, index=True)
    user_id = Column(Integer, ForeignKey("users.id"), nullable=True) # Optional: specific user or all if null
    type = Column(String) # 'ticket', 'vacation', etc.
    title = Column(String)
    message = Column(Text)
    is_read = Column(Integer, default=0) # 0 for false, 1 for true
    created_at = Column(DateTime, default=datetime.utcnow)

class RemarkIssue(Base):
    __tablename__ = "remark_issues"
    id = Column(Integer, primary_key=True, index=True)
    type = Column(String) # 'reported' or 'referenced'
    criticality = Column(String) # [A], [B], [C]
    issue_id = Column(String)
    description = Column(Text)
    team = Column(String)
    app_name = Column(String)
    created_at = Column(DateTime, default=datetime.utcnow)

class DailyProject(Base):
    __tablename__ = "daily_projects"
    id = Column(Integer, primary_key=True, index=True)
    name = Column(String, unique=True, index=True, nullable=False)
    created_at = Column(DateTime, default=datetime.utcnow)

class DailyIssue(Base):
    __tablename__ = "daily_issues"
    id = Column(Integer, primary_key=True, index=True)
    project_id = Column(Integer, ForeignKey("daily_projects.id"), nullable=False)
    title = Column(String, nullable=False)
    date = Column(String, nullable=False, index=True) # YYYY-MM-DD
    created_at = Column(DateTime, default=datetime.utcnow)

class ReferenceModel(Base):
    __tablename__ = "reference_models"
    id = Column(Integer, primary_key=True, index=True)
    model_name = Column(String, unique=True, index=True, nullable=False)
    ref_model = Column(Text, nullable=False) # Armazenado como texto pois pode ter múltiplos
    created_at = Column(DateTime, default=datetime.utcnow)

class STMSTranslation(Base):
    __tablename__ = "stms_translations"
    id = Column(Integer, primary_key=True, index=True)
    excel_row = Column(Integer)
    context = Column(Text)
    design_id = Column(Text)
    source_text = Column(Text)
    target_text = Column(Text)
    suggested_text = Column(Text)
    char_limit = Column(String)
    reason = Column(Text)
    simply_reason = Column(Text)
    status = Column(String, default='pending') # 'pending', 'reviewing', 'approved'
    source_filename = Column(Text)
    created_at = Column(DateTime, default=datetime.utcnow)
    design_type = Column(Text)

class TeamBoardArea(Base):
    __tablename__ = "team_board_areas"
    id = Column(Integer, primary_key=True, index=True)
    name = Column(String, nullable=False)
    section = Column(String, default="middle") # 'top', 'middle', 'bottom'
    position = Column(Integer, default=0)
    tab = Column(String, default="fixed") # 'fixed', 'current'
    created_at = Column(DateTime, default=datetime.utcnow)

class TeamBoardMember(Base):
    __tablename__ = "team_board_members"
    id = Column(Integer, primary_key=True, index=True)
    area_id = Column(Integer, ForeignKey("team_board_areas.id", ondelete="CASCADE"), nullable=False)
    user_id = Column(Integer, ForeignKey("users.id", ondelete="SET NULL"), nullable=True)
    name = Column(String, nullable=False)
    role = Column(String, nullable=True)
    identifier = Column(String, nullable=True)
    status = Column(String, default="normal") # 'normal', 'intern', 'movement', 'training'
    prefix = Column(String, nullable=True)
    parent_id = Column(Integer, ForeignKey("team_board_members.id", ondelete="CASCADE"), nullable=True)
    is_highlighted = Column(Integer, default=0)
    date_range = Column(String, nullable=True)
    position = Column(Integer, default=0)
    created_at = Column(DateTime, default=datetime.utcnow)

class TeamBoardProject(Base):
    __tablename__ = "team_board_projects"
    id = Column(Integer, primary_key=True, index=True)
    name = Column(String, nullable=False)
    position = Column(Integer, default=0)
    tab = Column(String, default="fixed") # 'fixed', 'current'
    created_at = Column(DateTime, default=datetime.utcnow)

class TeamBoardProjectMember(Base):
    __tablename__ = "team_board_project_members"
    id = Column(Integer, primary_key=True, index=True)
    project_id = Column(Integer, ForeignKey("team_board_projects.id", ondelete="CASCADE"), nullable=False)
    member_id = Column(Integer, ForeignKey("team_board_members.id", ondelete="CASCADE"), nullable=False)

class KanbanCard(Base):
    __tablename__ = "kanban_cards"
    id = Column(Integer, primary_key=True, index=True)
    title = Column(String, nullable=False)
    description = Column(Text, nullable=True)
    status = Column(String, default="Backlog")  # "Backlog", "To Do", "On Going", "Done"
    type = Column(String, default="pessoal")    # "pessoal", "time"
    user_id = Column(Integer, ForeignKey("users.id", ondelete="CASCADE"), nullable=True)
    project_id = Column(Integer, ForeignKey("team_board_projects.id", ondelete="SET NULL"), nullable=True)
    assigned_member_id = Column(Integer, ForeignKey("team_board_members.id", ondelete="SET NULL"), nullable=True)
    priority = Column(String, default="Média")
    position = Column(Integer, default=0)
    deadline = Column(String, nullable=True)
    created_at = Column(DateTime, default=datetime.utcnow)

class SystemNotice(Base):
    __tablename__ = "system_notices"
    id = Column(Integer, primary_key=True, index=True)
    title = Column(String, nullable=False)
    description = Column(Text, nullable=False)
    is_active = Column(Integer, default=0) # 0 for false, 1 for true
    created_at = Column(DateTime, default=datetime.utcnow)
    updated_at = Column(DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

# --- FUNÇÃO DE SINCRONIZAÇÃO AUTOMÁTICA ---
def sync_database():
    print("--- Iniciando Sincronização do Banco de Dados ---")
    Base.metadata.create_all(bind=engine)
    inspector = inspect(engine)
    
    # Mapeamento de colunas necessárias por tabela
    required_columns = {
        "users": {
            "cell": "TEXT",
            "kp": "TEXT",
            "kp_type": "TEXT",
            "is_backup": "INTEGER DEFAULT 0",
            "is_specialist": "INTEGER DEFAULT 0",
            "avatar": "TEXT",
            "bio": "TEXT",
            "skills": "TEXT"
        },
        "vacations": {
            "category": "TEXT DEFAULT 'vacation'",
            "sell_days": "INTEGER DEFAULT 0"
        },
        "stms_translations": {
            "simply_reason": "TEXT",
            "suggested_text": "TEXT",
            "design_id": "TEXT"
        },
        "team_board_members": {
            "user_id": "INTEGER",
            "position": "INTEGER DEFAULT 0"
        },
        "team_board_projects": {
            "position": "INTEGER DEFAULT 0",
            "tab": "TEXT DEFAULT 'fixed'"
        },
        "kanban_cards": {
            "title": "VARCHAR",
            "description": "TEXT",
            "status": "VARCHAR",
            "type": "VARCHAR",
            "user_id": "INTEGER",
            "project_id": "INTEGER",
            "assigned_member_id": "INTEGER",
            "priority": "VARCHAR DEFAULT 'Média'",
            "position": "INTEGER DEFAULT 0",
            "deadline": "VARCHAR"
        }
    }

    with engine.connect() as conn:
        for table_name, columns in required_columns.items():
            if table_name in inspector.get_table_names():
                existing_cols = {c['name'] for c in inspector.get_columns(table_name)}
                for col_name, col_type in columns.items():
                    if col_name not in existing_cols:
                        print(f"Sincronizando: Adicionando coluna '{col_name}' em '{table_name}'...")
                        try:
                            conn.execute(text(f"ALTER TABLE {table_name} ADD COLUMN {col_name} {col_type}"))
                        except Exception as e:
                            print(f"Aviso ao adicionar {col_name}: {e}")
        conn.commit()
    print("--- Banco de Dados Sincronizado com Sucesso ---")

sync_database()

app = FastAPI(title="SVP Server Integrado")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["GET", "POST", "PUT", "DELETE", "PATCH", "OPTIONS"],
    allow_headers=["*"],
)

# --- INTEGRAÇÃO DO STMS TOOL XML ---

# --- CONFIGURAÇÃO DA API DE INTELIGÊNCIA ARTIFICIAL (SIDIA PROXY) ---
# Endpoint do proxy corporativo OpenWebUI da Sidia
PROXY_URL = "https://openwebui.sidia.org.br/api/chat/completions"
# Chave de autenticação para o proxy da IA
API_KEY = 'sk-22424335d6864b07bf8f21ce0a20e5f4'
# Modelo de linguagem utilizado nas revisões e análises
MODEL_ID = "openai/gpt-oss-120b"

DB_CONFIG_XML = {
    "dbname": "stms_tool_xml",
    "user": "postgres",
    "password": "326741@Vv",
    "host": "localhost",
    "port": "5432"
} 
 
try:
    db_pool_xml = pool.SimpleConnectionPool(1, 40, **DB_CONFIG_XML)
except Exception as e:
    print(f"Aviso - Não foi possível criar o pool de banco de dados XML: {e}")
    db_pool_xml = None

def init_db_xml():
    try:
        conn = psycopg2.connect(**DB_CONFIG_XML)
        cur = conn.cursor()
        create_table_query = """
        CREATE TABLE IF NOT EXISTS glossary (
            no_id TEXT,
            english TEXT,
            language TEXT,
            translation TEXT,
            description TEXT,
            parts_of_speech TEXT,
            polysemy TEXT,
            dnt TEXT,
            app_name TEXT,
            abbreviation TEXT,
            tag TEXT,
            trans_status TEXT,
            language_context TEXT,
            id TEXT PRIMARY KEY
        );
        CREATE TABLE IF NOT EXISTS glossary_es (
            no_id TEXT,
            english TEXT,
            language TEXT,
            translation TEXT,
            description TEXT,
            parts_of_speech TEXT,
            polysemy TEXT,
            dnt TEXT,
            app_name TEXT,
            abbreviation TEXT,
            tag TEXT,
            trans_status TEXT,
            language_context TEXT,
            id TEXT PRIMARY KEY
        );
        """
        cur.execute(create_table_query)
        conn.commit()
        cur.close()
        conn.close()
        print("Banco de dados XML verificado com sucesso.")
    except Exception as e:
        print(f"Aviso - Não foi possível conectar ao banco de dados XML: {e}")

def search_glossary(english_text, target_lang="pt"):
    """Busca exata — retorna um único match se o texto inteiro for um termo do glossário."""
    if not db_pool_xml: return None
    conn = None
    try:
        conn = db_pool_xml.getconn()
        cur = conn.cursor()
        table = "glossary_es" if target_lang == "es" else "glossary"
        query = f"SELECT translation, description, dnt FROM {table} WHERE LOWER(english) = LOWER(%s) LIMIT 1"
        cur.execute(query, (english_text.strip(),))
        result = cur.fetchone()
        cur.close()
        return result
    except Exception as e:
        return None
    finally:
        if conn:
            db_pool_xml.putconn(conn)

def search_glossary_multi(english_text, target_lang="pt"):
    """Busca inteligente — encontra TODOS os termos do glossário que aparecem dentro do texto fonte."""
    if not db_pool_xml: return []
    conn = None
    try:
        conn = db_pool_xml.getconn()
        cur = conn.cursor()
        # Busca todos os termos do glossário que estão contidos no texto (case-insensitive)
        table = "glossary_es" if target_lang == "es" else "glossary"
        cur.execute(f"SELECT english, translation, description, dnt, app_name FROM {table}")
        all_terms = cur.fetchall()
        cur.close()
        
        matches = []
        text_lower = english_text.lower()
        for term in all_terms:
            term_english = (term[0] or "").strip()
            if not term_english:
                continue
            # Verifica se o termo do glossário aparece como palavra inteira no texto
            pattern = r'\b' + re.escape(term_english.lower()) + r'\b'
            if re.search(pattern, text_lower):
                matches.append({
                    "english": term_english,
                    "translation": term[1] or "",
                    "description": term[2] or "",
                    "dnt": term[3] or "No",
                    "app_name": term[4] or ""
                })
        return matches
    except Exception as e:
        return []
    finally:
        if conn:
            db_pool_xml.putconn(conn)

init_db_xml()

def parse_xml(content):
    res = {}
    if not content or not isinstance(content, str): return res
    try:
        c = content.strip()
        if "<!doctype html>" in c[:500].lower() or "gitiles" in c[:500].lower():
            lines = re.findall(r'<td class="FileContents-lineContents".*?>(.*?)</td>', c)
            if lines:
                decoded_lines = []
                for line in lines:
                    clean_line = re.sub(r'<[^>]+>', '', line)
                    decoded_lines.append(html.unescape(clean_line))
                c = '\n'.join(decoded_lines)

        m = re.search(r'<resources.*?>.*</resources>', c, re.DOTALL)
        if m: c = m.group(0)
        else:
            m2 = re.search(r'<resources.*?>', c)
            if m2: c = c[m2.start():]
        comment_map = {}
        for match in re.finditer(r'<!--(.*?)-->\s*<[^>]+name="([^"]+)"', c, re.DOTALL):
            comment_text = match.group(1).strip()
            string_name = match.group(2)
            comment_map[string_name] = comment_text
            
        c = re.sub(r'<!--.*?-->', '', c, flags=re.DOTALL)
        root = ET.fromstring(c)
        for el in root:
            tag = el.tag.split('}')[-1]
            name = el.get('name')
            if not name: continue
            
            comment = comment_map.get(name, "")
            
            if tag == 'string': res[name] = {"value": (el.text or "").strip(), "comment": comment}
            elif tag == 'plurals':
                for i in el.findall('item'):
                    if i.get('quantity') == 'other':
                        res[name] = {"value": (i.text or "").strip(), "comment": comment}
                        break
            elif tag == 'string-array':
                for i, it in enumerate(el.findall('item')):
                    res[f"{name}[{i}]"] = {"value": (it.text or "").strip(), "comment": comment}
    except Exception as e: 
        print("Parse error:", e)
    return res

def get_tov_sections(target_lang="pt"):
    filename = "tone_of_voice_es.txt" if target_lang == "es" else "tone_of_voice.txt"
    kb_path = os.path.join(os.path.dirname(__file__), 'knowledge_base', filename)
    if not os.path.exists(kb_path):
        return {}, ""
    
    try:
        with open(kb_path, 'r', encoding='utf-8') as f:
            content = f.read()
            
        sections = {}
        current_header = "GENERAL"
        current_lines = []
        
        for line in content.split('\n'):
            if line.startswith('#'):
                if current_lines:
                    sections[current_header] = '\n'.join(current_lines).strip()
                current_header = line.strip()
                current_lines = []
            else:
                current_lines.append(line)
        
        if current_lines:
            sections[current_header] = '\n'.join(current_lines).strip()
            
        return sections, content
    except Exception as e:
        print(f"Erro ao ler tone_of_voice.txt: {e}")
        return {}, ""

def get_dynamic_rules(key, en_content, pt_content, en_comment, pt_comment, target_lang="pt"):
    tov_sections, tov_raw = get_tov_sections(target_lang)
    design_type = "GERAL"
    
    if not tov_sections:
        return "Responda de forma profissional e direta.", design_type

    # --- INJEÇÃO INTELIGENTE DA BASE DE CONHECIMENTO ---
    # Somente insere os trechos do arquivo tone_of_voice.txt que realmente importam para essa string específica
    relevant_rules = ""
    
    def add_section_by_keyword(keywords, add_all_matches=False):
        added = False
        for k, text in tov_sections.items():
            if any(kw.lower() in k.lower() for kw in keywords):
                nonlocal relevant_rules
                relevant_rules += f"{k}\n{text}\n\n"
                added = True
                if not add_all_matches: break
        return added

    # 1. Regras Base (Sempre adicionar as características e estilos)
    add_section_by_keyword(["tom de voz", "idioma oficial", "estilo de localização"])
    add_section_by_keyword(["gramática", "maiúsculas"], add_all_matches=True)
    add_section_by_keyword(["pontuação"], add_all_matches=True)

    # 2. Match Específico por Design Type (Aba, Pop-up, Botão, etc)
    for k, text in tov_sections.items():
        if "(HEADER)" in k or "(BUTTON)" in k or "(OPT" in k or "(BODY" in k or "(STATUS)" in k or "(NPBODY)" in k:
            tags = re.findall(r'\(([A-Z/]+)\)', k)
            if tags:
                matched = False
                for tag_group in tags:
                    for t in tag_group.split('/'):
                        if t in en_comment or t in pt_comment or f"_{t}_" in key.upper() or key.upper().startswith(t):
                            relevant_rules += f"[REGRA ESPECÍFICA DO DESIGN ({t})]\n{k}\n{text}\n\n"
                            design_type = t
                            matched = True
                            break
                    if matched: break
    
    # 3. Match Dinâmico por Expressões (Se tiver número, chama a regra de número, etc)
    if any(s in en_content for s in ['+', '-', 'x', '÷', '=', '>', '<', '%', '~', '.']):
        add_section_by_keyword(["espaço", "símbolo"], add_all_matches=True)
        
    if re.search(r'\d', en_content) or re.search(r'\b(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec|Mon|Tue|Wed|Thu|Fri|Sat|Sun)\b', en_content, re.IGNORECASE):
        add_section_by_keyword(["data", "hora", "número", "unidade", "porcentagem"], add_all_matches=True)
        
    if re.search(r'%(s|d|f|\d\$[sdf])', en_content):
        add_section_by_keyword(["reservado"], add_all_matches=True)
        
    if 'TTS' in en_comment or 'TTS' in key.upper() or 'DREAM_SAY' in key.upper():
        add_section_by_keyword(["tts", "leitura"], add_all_matches=True)
        
    if en_content.startswith('S ') or 'Galaxy' in en_content or 'SmartThings' in en_content or 'SideSync' in en_content or 'Pay' in en_content:
        add_section_by_keyword(["aplicativo", "dnt", "não traduza", "recursos"], add_all_matches=True)
        
    dei_keywords = ['old', 'age', 'year', 'disable', 'blind', 'deaf', 'see', 'hear', 'family', 'child', 'son', 'daughter', 'gender', 'sex', 'he', 'she', 'his', 'her']
    if any(re.search(rf'\b{kw}\b', en_content.lower()) for kw in dei_keywords):
        add_section_by_keyword(["dei", "princípio"], add_all_matches=True)
        
    if re.search(r'\b[A-Za-zÀ-ÿ0-9_]{2,}\.', pt_content):
        add_section_by_keyword(["abreviação"], add_all_matches=True)

    # Injetar feedback de erros anteriores reportados pelo usuário
    try:
        feedback_filename = f"feedback_{target_lang}.txt"
        feedback_path = os.path.join(os.path.dirname(__file__), 'knowledge_base', feedback_filename)
        if os.path.exists(feedback_path):
            with open(feedback_path, 'r', encoding='utf-8') as f:
                feedback_content = f.read().strip()
            if feedback_content:
                # Limitar para os últimos 50 feedbacks para não sobrecarregar o prompt
                feedback_entries = feedback_content.split("--- FEEDBACK")
                recent_entries = feedback_entries[-50:] if len(feedback_entries) > 50 else feedback_entries
                trimmed_feedback = "--- FEEDBACK".join(recent_entries).strip()
                relevant_rules += f"\n\n[ERROS ANTERIORES REPORTADOS PELO USUÁRIO — NÃO REPITA ESTES ERROS:]\n{trimmed_feedback}\n"
    except Exception as e:
        print(f"Aviso: Não foi possível carregar feedback: {e}")

    relevant_rules += "\n\n[REGRA DE REDUNDÂNCIA]\nNÃO corrija ou relate redundâncias. Se a tradução alvo parecer redundante, IGNORE. A redundância deve ser mantida como correta. Nenhuma alteração deve ser feita por motivo de redundância.\n"

    return relevant_rules.strip(), design_type

# --- INTEGRAÇÃO COM A API DA IA (Não altere a lógica) ---
def fetch_translation(item, target_lang="pt"):
    key = item['string_name']
    en_content = item['en']
    pt_content = item['pt']
    en_comment = item.get('en_comment', '').upper()
    pt_comment = item.get('pt_comment', '').upper()
    
    glossary_match = search_glossary(en_content, target_lang)
    glossary_hint = ""
    if glossary_match:
        trans, desc, dnt = glossary_match
        if dnt == 'Y':
            glossary_hint = f"\n[REGRA DE OURO DO GLOSSÁRIO: Manter como '{en_content}' (DNT)]"
        else:
            glossary_hint = f"\n[REGRA DE OURO DO GLOSSÁRIO: Tradução Obrigatória: '{trans}'. Contexto: {desc}]"
    
    # Busca inteligente: encontra TODOS os termos do glossário contidos no texto
    glossary_multi = search_glossary_multi(en_content, target_lang)
    if glossary_multi:
        glossary_rules = []
        for g in glossary_multi:
            if g['dnt'] in ('Y', 'Yes', 'y', 'yes'):
                glossary_rules.append(f"- '{g['english']}': NÃO TRADUZIR (DNT). Manter exatamente como está.")
            else:
                glossary_rules.append(f"- '{g['english']}' → Traduzir como '{g['translation']}'. {g['description']}")
        glossary_hint += "\n[TERMOS DO GLOSSÁRIO ENCONTRADOS NO TEXTO — RESPEITE CADA UM:]\n" + "\n".join(glossary_rules)

    relevant_rules, design_type = get_dynamic_rules(key, en_content, pt_content, en_comment, pt_comment, target_lang)

    headers = {
        "Authorization": f"Bearer {API_KEY}",
        "Content-Type": "application/json"
    }
    
    lang_name = "ESPANHOL" if target_lang == "es" else "PORTUGUÊS"
    lang_code = "ES" if target_lang == "es" else "PT"
    
    system_prompt = f"""Você é o Revisor de Tradução da SAMSUNG.
DIRETRIZES APLICÁVEIS PARA ESTA STRING:
{relevant_rules}

INSTRUÇÕES (LEIA COM ATENÇÃO):
- Analise SEVERAMENTE o texto '{lang_code}' atual. Ele atende ao Tom de Voz da Samsung, regras gramaticais e de design ({design_type}) listadas acima?
- Se o '{lang_code}' violar QUALQUER regra (ex: falta de espaço na unidade, erro gramatical), corrija-o obrigatoriamente.
- Se a Regra de Ouro do Glossário estiver presente, ELA É SOBERANA.
- Responda EXATAMENTE neste formato XML: 
<advice>sugestão corrigida ou 'Mantido' se estiver perfeito</advice>
<reason>motivo detalhado da alteração baseada na regra. IMPORTANTE: Escreva sempre em Português (PT-BR).</reason>
<simplyReason>resumo curto do erro (ex: 'Falta de espaço'). Retorne 'Correto' SOMENTE SE advice for 'Mantido'. IMPORTANTE: Escreva sempre em Português (PT-BR).</simplyReason>
Revise o texto priorizando a fidelidade ao original em inglês. Mantenha o texto o mais próximo possível do original, removendo apenas redundâncias óbvias e corrigindo erros graves de pontuação, gramática ou formatação. Preserve termos técnicos, jargões e abreviações, e evite adicionar palavras ou sinônimos que alterem o significado. Corrija apenas o necessário para garantir clareza e fidelidade ao original.
"""

    user_content = f"Chave: {key} | Contexto: {design_type}\nEN: {en_content}\n{lang_code}: {pt_content}{glossary_hint}"
    
    messages = [
        {"role": "system", "content": system_prompt},
        {"role": "user", "content": user_content}
    ]

    payload = {
        "model": MODEL_ID,
        "messages": messages,
        "stream": False
    }
    
    max_retries = 3
    for attempt in range(max_retries):
        try:
            response = requests.post(PROXY_URL, json=payload, headers=headers, timeout=200, verify=False)
            if response.status_code == 200:
                res_data = response.json()
                content = res_data.get('message', {}).get('content')
                if not content and 'choices' in res_data and len(res_data['choices']) > 0:
                    content = res_data['choices'][0].get('message', {}).get('content', '')
                if not content:
                    content = ''
                
                match = re.search(r'<advice>(.*?)</advice>.*?<reason>(.*?)</reason>.*?<simplyReason>(.*?)</simplyReason>', content, re.DOTALL)
                advice = match.group(1).strip() if match else "Sem sugestão"
                reason = match.group(2).strip() if match else "Tradução OK"
                simply = match.group(3).strip() if match else reason
                
                # Trava de segurança no Python: se alterou algo (inclusive uma vírgula), NÃO pode ser 'Correto'
                if advice != "Mantido" and advice != pt_content:
                    if simply.lower() in ["correto", "preciso", "ok", "perfeito"]:
                        simply = "Alteração de formatação/pontuação"
                
                return {**item, "advice": advice, "reason": reason, "simplyReason": simply}
            else:
                if attempt == max_retries - 1:
                    return {**item, "advice": "ERRO", "reason": f"HTTP {response.status_code}", "simplyReason": "Erro API"}
                time.sleep(2 * (attempt + 1))
        except Exception as e:
            if attempt == max_retries - 1:
                return {**item, "advice": "ERRO", "reason": str(e), "simplyReason": "Exception"}
            time.sleep(2 * (attempt + 1))

@app.post("/parse")
async def parse_files(files: List[UploadFile] = File(...)):
    en_dicts = {}
    pt_dicts = {}
    app_names = {}
    idx_set = set()
    
    for f in files:
        content_bytes = await f.read()
        content = content_bytes.decode('utf-8', errors='ignore')
        name = f.filename.lower().replace('.xml', '')
        parts = name.split('_')
        
        if len(parts) >= 3:
            lang = parts[-2]
            idx = parts[-1]
            app_prefix = "_".join(parts[:-2])
            key = f"{app_prefix}_{idx}"
            
            app_names[key] = app_prefix
            
            parsed_dict = parse_xml(content)
            if lang == 'en':
                en_dicts[key] = parsed_dict
            elif lang in ['pt', 'br', 'pt-br']:
                pt_dicts[key] = parsed_dict
            idx_set.add(key)
            
    merged_items = []
    for k in idx_set:
        if k in en_dicts and k in pt_dicts:
            en_d = en_dicts[k]
            pt_d = pt_dicts[k]
            app_name = app_names[k]
            
            keys = set(en_d.keys()) | set(pt_d.keys())
            for string_key in keys:
                en_item = en_d.get(string_key, {"value": "None", "comment": ""})
                pt_item = pt_d.get(string_key, {"value": "None", "comment": ""})
                
                merged_items.append({
                    "app_name": app_name,
                    "string_name": string_key,
                    "en": en_item.get("value", "None"),
                    "en_comment": en_item.get("comment", ""),
                    "pt": pt_item.get("value", "None"),
                    "pt_comment": pt_item.get("comment", "")
                })
                
    return {"items": merged_items}
class SystemNoticeCreate(BaseModel):
    title: str
    description: str
    is_active: bool = False

class SystemNoticeUpdate(BaseModel):
    title: Optional[str] = None
    description: Optional[str] = None
    is_active: Optional[bool] = None


class BatchRequest(BaseModel):
    items: List[Dict[str, Any]]
    target_lang: Optional[str] = "pt"

@app.post("/process_batch")
async def process_batch(request: BatchRequest):
    items = request.items
    results = []
    
    with concurrent.futures.ThreadPoolExecutor(max_workers=40) as executor:
        futures = [executor.submit(fetch_translation, item, request.target_lang) for item in items]
        for future in concurrent.futures.as_completed(futures):
            results.append(future.result())
            
    return {"results": results}

class ReportRequest(BaseModel):
    projectInfo: Dict[str, Any]
    items: List[Dict[str, Any]]

@app.post("/generate_report")
async def generate_report(request: ReportRequest):
    project_info = request.projectInfo
    items = request.items
    
    df = pd.DataFrame(items)
    df_excel = df[['app_name', 'string_name', 'en', 'pt', 'advice', 'reason', 'simplyReason']].copy()
    df_excel.columns = ['App', 'Chave', 'Inglês Original', 'Português Original', 'Sugestão IA', 'Motivo Técnico', 'Resumo']
    
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_excel.to_excel(writer, index=False, sheet_name='Traduções')
        
        workbook = writer.book
        worksheet = writer.sheets['Traduções']
        red_font = InlineFont(color='FFFF0000')
        
        for row in range(2, len(df_excel) + 2):
            orig_cell = worksheet.cell(row=row, column=4)
            sugg_cell = worksheet.cell(row=row, column=5)
            
            original = str(orig_cell.value) if orig_cell.value else ""
            suggestion = str(sugg_cell.value) if sugg_cell.value else ""
            
            if suggestion and suggestion != "Mantido" and original != suggestion:
                matcher = difflib.SequenceMatcher(None, original, suggestion)
                rich_text_elements = []
                
                for tag, i1, i2, j1, j2 in matcher.get_opcodes():
                    text_part = suggestion[j1:j2]
                    if not text_part:
                        continue
                    
                    if tag == 'equal':
                        rich_text_elements.append(text_part)
                    elif tag in ('insert', 'replace'):
                        rich_text_elements.append(TextBlock(font=red_font, text=text_part))
                
                has_changes = any(isinstance(x, TextBlock) for x in rich_text_elements)
                if has_changes:
                    sugg_cell.value = CellRichText(rich_text_elements)
    
    excel_data = base64.b64encode(output.getvalue()).decode()

    headers = {
        "Authorization": f"Bearer {API_KEY}",
        "Content-Type": "application/json"
    }
    
    email_prompt = f"""
    Escreva um e-mail profissional EM INGLÊS informando que a revisão das traduções para o projeto '{project_info.get('refObjectName')}' (Versão: {project_info.get('swVersion')}, Feature: {project_info.get('feature')}) foi concluída e o relatório está em anexo.
    REGRAS CRÍTICAS: (TUDO EM INGLÊS) - Não precisa colocar [Recipient], APENAS DEAR,
    - NÃO use caracteres especiais de formatação como asteriscos (*), hifens iniciais (-), sublinhados (_) ou tils (~).
    - Use apenas quebras de linha simples e texto puro.
    - Deve explicar: que a revisão das traduções para o projeto '{project_info.get('refObjectName')}' (Versão: {project_info.get('swVersion')}, Feature: {project_info.get('feature')}) foi concluída e o relatório está em anexo. Colocando o que foi testado, nome do aplicativo e funções: '{project_info.get('refObjectName')}' (Versão: {project_info.get('swVersion')}, Feature: {project_info.get('feature')})
    - Deixe claro que o texto do email foi gerado por IA no SIDIA.
    - O tom deve ser profissional e executivo.
    - Mencione que o arquivo Excel com os detalhes está anexado.
    - Finalize com BR, STMS Automation Team.
    """

    messages = [
        {"role": "system", "content": "Você é um assistente corporativo que escreve e-mails profissionais em texto puro, sem qualquer formatação Markdown ou caracteres especiais."},
        {"role": "user", "content": email_prompt}
    ]

    payload = {
        "model": MODEL_ID,
        "messages": messages,
        "stream": False
    }

    email_body = "Dear all, please find attached the revised translation report for the project reviewed by AI."
    try:
        response = requests.post(PROXY_URL, json=payload, headers=headers, timeout=20, verify=False)
        if response.status_code == 200:
            res_data = response.json()
            content = res_data.get('message', {}).get('content')
            if not content and 'choices' in res_data and len(res_data['choices']) > 0:
                content = res_data['choices'][0].get('message', {}).get('content')
            
            if content:
                email_body = content
            email_body = re.sub(r'[*_~`#\-]', '', email_body)
    except:
        pass

    return {
        "to": ["gilmar.silva@samsung.com ;", "edgard.cunha@samsung.com ;"],
        "cc": ["ivan.moreira@samsung.com ;", "wallid.m@samsung.com"],
        "subject": f"AI Translation Report - {project_info.get('refObjectName')}",
        "body": email_body,
        "excel_base64": excel_data,
        "filename": f"Relatorio_Traducao_{project_info.get('refObjectName').replace(' ', '_')}.xlsx"
    }

# --- SCHEMAS ORIGINAIS ---

class LoginRequest(BaseModel):
    email: str
    password: str

class FirstAccessRequest(BaseModel):
    email: str
    display_name: Optional[str] = None
    team: str
    cell: Optional[str] = None
    kp: str
    role: str
    sidia_id: str
    kp_type: Optional[str] = None
    is_backup: Optional[bool] = False
    is_specialist: Optional[int] = 0

class RemarkData(BaseModel):
    team: str = "Unknown"
    testerId: Optional[str] = ""
    account: Optional[str] = ""
    samsungAccount: Optional[str] = ""
    simCard: Optional[str] = ""
    appName: Optional[str] = "" 
    sampleId: Optional[str] = ""
    deviceId: Optional[str] = ""
    testerIds: Optional[List[str]] = []
    accounts: Optional[List[str]] = []
    samsungAccounts: Optional[List[str]] = []
    simCards: Optional[List[str]] = []
    appNames: Optional[List[str]] = []
    sampleIds: Optional[List[str]] = []
    deviceIds: Optional[List[str]] = []
    full_form_data: Optional[Dict[str, Any]] = {}

class VacationPeriod(BaseModel):
    start: str
    end: str
    category: Optional[str] = "vacation"
    status: Optional[str] = "pending"

class VacationRequest(BaseModel):
    userId: int
    periods: List[VacationPeriod]
    sellDays: Optional[bool] = False

class VacationStatusUpdate(BaseModel):
    status: str

class TicketStatusUpdate(BaseModel):
    status: Optional[str] = None
    resolution: Optional[str] = None

class ReferenceModelBase(BaseModel):
    model_name: str
    ref_model: str

class ReferenceModelCreate(ReferenceModelBase):
    pass

class ReferenceModelUpdate(ReferenceModelBase):
    pass

class TicketCreate(BaseModel):
    type: str
    priority: str
    status: Optional[str] = "pendente"
    title: str
    content: str
    resolution: Optional[str] = None
    creators: List[Dict[str, str]]

class DailyProjectCreate(BaseModel):
    name: str

class DailyIssueCreate(BaseModel):
    projectId: int
    title: str
    date: str

class STMSReviewRequest(BaseModel):
    id: int
    context: Optional[str] = ""
    source_text: str
    target_text: Optional[str] = ""
    char_limit: Optional[str] = ""
    target_lang: Optional[str] = "pt"

class STMSApproveRequest(BaseModel):
    id: int

class TeamBoardAreaCreate(BaseModel):
    name: str
    section: str = "middle"
    position: int = 0
    tab: str = "fixed"

class TeamBoardMemberCreate(BaseModel):
    area_id: int
    name: str
    role: Optional[str] = None
    identifier: Optional[str] = None
    status: str = "normal"
    prefix: Optional[str] = None
    parent_id: Optional[int] = None
    user_id: Optional[int] = None
    is_highlighted: Optional[int] = 0
    date_range: Optional[str] = None
    position: Optional[int] = 0

def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

def create_notification(db: Session, title: str, message: str, type: str, user_id: Optional[int] = None):
    notif = Notification(
        title=title,
        message=message,
        type=type,
        user_id=user_id
    )
    db.add(notif)
    db.commit()
    db.refresh(notif)
    return notif

# --- ROTAS STMS TOOL ---

@app.get("/stms/strings")
def get_stms_strings(db: Session = Depends(get_db)):
    return db.query(STMSTranslation).order_by(STMSTranslation.excel_row.asc()).all()

@app.post("/stms/review_ai")
def review_stms_ai(item: STMSReviewRequest, db: Session = Depends(get_db)):
    # Mapear para o formato que fetch_translation entende
    # fetch_translation(item) espera: string_name, en, pt, en_comment, pt_comment
    
    mapping_item = {
        "string_name": f"ROW_{item.id}",
        "en": item.source_text,
        "pt": item.target_text or "",
        "en_comment": item.context or "",
        "pt_comment": ""
    }
    
    # Adicionar dica de limite de caracteres se existir
    if item.char_limit:
        mapping_item["en_comment"] += f" (MAX: {item.char_limit})"
        
    try:
        # Chama a lógica principal de revisão da IA que já existe no servidor
        result = fetch_translation(mapping_item, target_lang=item.target_lang)
        
        advice = result.get('advice', item.target_text)
        reason = result.get('reason', 'Revisão automática realizada.')
        simply = result.get('simplyReason', 'Revisão OK')
        
        # Se a IA retornou erro, NÃO salvar no banco — devolver o erro ao frontend
        if advice == "ERRO":
            return {"suggestion": item.target_text, "reasoning": f"Erro na API da IA: {reason}", "simplyReason": simply, "error": True}
        
        if advice == "Mantido":
            advice = item.target_text
            
        # Atualiza Banco somente com resultados válidos
        db_item = db.query(STMSTranslation).filter(STMSTranslation.id == item.id).first()
        if db_item:
            db_item.suggested_text = advice
            db_item.reason = reason
            db_item.simply_reason = simply
            db_item.status = 'reviewing'
            db.commit()
            
        return {"suggestion": advice, "reasoning": reason, "simplyReason": simply}
    except Exception as e:
        print(f"Erro na revisão STMS: {e}")
        return {"suggestion": item.target_text, "reasoning": f"Erro Local: {str(e)}", "error": True}

@app.post("/stms/approve_string")
def approve_stms_string(data: STMSApproveRequest, db: Session = Depends(get_db)):
    try:
        db_item = db.query(STMSTranslation).filter(STMSTranslation.id == data.id).first()
        if db_item:
            db_item.status = 'approved'
            db.commit()
            return {"status": "success"}
        raise HTTPException(status_code=404, detail="Item não encontrado")
    except Exception as e:
        db.rollback()
        print(f"Erro ao aprovar string {data.id}: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/stms/reject_string")
def reject_stms_string(data: STMSApproveRequest, db: Session = Depends(get_db)):
    try:
        db_item = db.query(STMSTranslation).filter(STMSTranslation.id == data.id).first()
        if db_item:
            db_item.status = 'rejected'
            db_item.suggested_text = db_item.target_text
            db.commit()
            return {"status": "success", "suggestion": db_item.target_text}
        raise HTTPException(status_code=404, detail="Item não encontrado")
    except Exception as e:
        db.rollback()
        print(f"Erro ao rejeitar string {data.id}: {e}")
        raise HTTPException(status_code=500, detail=str(e))

class STMSFeedbackRequest(BaseModel):
    id: int
    feedback: str
    ai_reason: Optional[str] = ""
    source_text: Optional[str] = ""
    target_text: Optional[str] = ""
    suggested_text: Optional[str] = ""
    target_lang: Optional[str] = "pt"
    action: Optional[str] = "reject"  # "reject" or "re_review"

@app.post("/stms/feedback")
def submit_stms_feedback(data: STMSFeedbackRequest, db: Session = Depends(get_db)):
    """Recebe feedback do usuário sobre erro da IA e salva na base de conhecimento."""
    try:
        # 1. Salvar feedback no arquivo de conhecimento
        feedback_filename = f"feedback_{data.target_lang}.txt"
        kb_dir = os.path.join(os.path.dirname(__file__), "knowledge_base")
        os.makedirs(kb_dir, exist_ok=True)
        feedback_path = os.path.join(kb_dir, feedback_filename)
        
        from datetime import datetime
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M")
        
        feedback_entry = f"\n--- FEEDBACK ({timestamp}) ---\n"
        feedback_entry += f"EN: {data.source_text}\n"
        lang_code = "ES" if data.target_lang == "es" else "PT"
        feedback_entry += f"{lang_code}: {data.target_text}\n"
        feedback_entry += f"Sugestão da IA: {data.suggested_text}\n"
        feedback_entry += f"Motivo da IA: {data.ai_reason}\n"
        feedback_entry += f"ERRO REPORTADO PELO USUÁRIO: {data.feedback}\n"
        feedback_entry += f"---\n"
        
        with open(feedback_path, "a", encoding="utf-8") as f:
            f.write(feedback_entry)
        
        print(f"[FEEDBACK] Salvo em {feedback_filename}: {data.feedback[:80]}...")
        
        # 2. Executar a ação (rejeitar ou re-revisar)
        db_item = db.query(STMSTranslation).filter(STMSTranslation.id == data.id).first()
        if not db_item:
            raise HTTPException(status_code=404, detail="Item não encontrado")
        
        if data.action == "reject":
            db_item.status = "rejected"
            db.commit()
            return {"status": "success", "action": "rejected"}
        
        elif data.action == "re_review":
            # Re-revisar: chamar a IA novamente com o feedback já salvo na base
            mapping_item = {
                "string_name": f"ROW_{data.id}",
                "en": data.source_text or db_item.source_text,
                "pt": data.target_text or db_item.target_text,
                "en_comment": db_item.context or "",
                "pt_comment": ""
            }
            
            if db_item.char_limit:
                mapping_item["en_comment"] += f" (MAX: {db_item.char_limit})"
            
            result = fetch_translation(mapping_item, target_lang=data.target_lang)
            
            advice = result.get('advice', db_item.target_text)
            reason = result.get('reason', 'Revisão automática realizada.')
            simply = result.get('simplyReason', 'Revisão OK')
            
            if advice == "ERRO":
                return {"status": "error", "action": "re_review", "suggestion": db_item.target_text, "reasoning": f"Erro na API: {reason}", "simplyReason": simply}
            
            if advice == "Mantido":
                advice = db_item.target_text
            
            db_item.suggested_text = advice
            db_item.reason = reason
            db_item.simply_reason = simply
            db_item.status = "reviewing"
            db.commit()
            
            return {
                "status": "success",
                "action": "re_reviewed",
                "suggestion": advice,
                "reasoning": reason,
                "simplyReason": simply
            }
        
        return {"status": "error", "detail": "Ação inválida"}
        
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        print(f"Erro ao processar feedback {data.id}: {e}")
        raise HTTPException(status_code=500, detail=str(e))

class STMSTranslateRequest(BaseModel):
    text: str
    target_lang: str

@app.post("/stms/translate_text")
def translate_stms_text(data: STMSTranslateRequest):
    try:
        dest = "Espanhol" if data.target_lang == "es" else "Português do Brasil (PT-BR)"
        prompt = f"Traduza o seguinte texto técnico para {dest}. Retorne APENAS a tradução, sem aspas ou comentários:\n\n{data.text}"
        
        response = chat_session.send_message(prompt)
        translation = response.text.strip()
        
        return {"translated_text": translation}
    except Exception as e:
        print(f"Erro ao traduzir texto: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/stms/postpone_string")
def postpone_stms_string(data: STMSApproveRequest, db: Session = Depends(get_db)):
    try:
        db_item = db.query(STMSTranslation).filter(STMSTranslation.id == data.id).first()
        if db_item:
            db_item.status = 'postponed'
            db.commit()
            return {"status": "success"}
        raise HTTPException(status_code=404, detail="Item não encontrado")
    except Exception as e:
        db.rollback()
        print(f"Erro ao adiar string {data.id}: {e}")
        raise HTTPException(status_code=500, detail=str(e))

class STMSUpdateRequest(BaseModel):
    id: int
    suggested_text: Optional[str] = None
    design_id: Optional[str] = None

@app.post("/stms/update_string")
def update_stms_string(data: STMSUpdateRequest, db: Session = Depends(get_db)):
    try:
        db_item = db.query(STMSTranslation).filter(STMSTranslation.id == data.id).first()
        if db_item:
            if data.suggested_text is not None:
                db_item.suggested_text = data.suggested_text
            if data.design_id is not None:
                db_item.design_id = data.design_id
            db.commit()
            return {"status": "success"}
        raise HTTPException(status_code=404, detail="Item não encontrado")
    except Exception as e:
        db.rollback()
        print(f"Erro ao atualizar string {data.id}: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/stms/undo_string")
def undo_stms_string(data: STMSApproveRequest, db: Session = Depends(get_db)):
    try:
        db_item = db.query(STMSTranslation).filter(STMSTranslation.id == data.id).first()
        if db_item:
            db_item.status = 'reviewing'
            db.commit()
            return {"status": "success"}
        raise HTTPException(status_code=404, detail="Item não encontrado")
    except Exception as e:
        db.rollback()
        print(f"Erro ao desfazer string {data.id}: {e}")
        raise HTTPException(status_code=500, detail=str(e))

class STMSDeleteFileRequest(BaseModel):
    filename: str

@app.post("/stms/delete_file")
def delete_stms_file(data: STMSDeleteFileRequest, db: Session = Depends(get_db)):
    try:
        # Deleta todas as strings associadas ao arquivo
        result = db.query(STMSTranslation).filter(STMSTranslation.source_filename == data.filename).delete()
        db.commit()
        return {"status": "success", "deleted_count": result}
    except Exception as e:
        db.rollback()
        print(f"Erro ao deletar arquivo {data.filename}: {e}")
        raise HTTPException(status_code=500, detail=str(e))

class STMSExportRequest(BaseModel):
    items: List[Dict[str, Any]]
    
@app.post("/stms/export_excel")
async def export_stms_excel(request: STMSExportRequest):
    items = request.items
    df = pd.DataFrame(items)
    
    # Selecionar e renomear colunas
    df_excel = df[['excel_row', 'source_filename', 'context', 'char_limit', 'source_text', 'target_text', 'suggested_text', 'simply_reason', 'reason', 'status']].copy()
    df_excel.columns = ['Row', 'File', 'Context', 'Limit', 'Source (EN)', 'Target (PT)', 'Suggested', 'Analysis', 'Full Reason', 'Status']
    
    # Mapear status e ajustar Target (PT) para refletir a decisão
    status_map = {
        'approved': 'APROVADO',
        'rejected': 'REJEITADO',
        'reviewing': 'REVISANDO',
        'pending': 'PENDENTE'
    }
    
    for i, row in df_excel.iterrows():
        orig_status = row['Status']
        df_excel.at[i, 'Status'] = status_map.get(orig_status, orig_status.upper())
        
        # Se aprovado, o Target (PT) final é a sugestão da IA
        if orig_status == 'approved':
            df_excel.at[i, 'Target (PT)'] = row['Suggested']
    
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_excel.to_excel(writer, index=False, sheet_name='DB Export')
        
        workbook = writer.book
        worksheet = writer.sheets['DB Export']
        red_font = InlineFont(color='FFFF0000')
        
        for row in range(2, len(df_excel) + 2):
            orig_cell = worksheet.cell(row=row, column=6) # Target
            sugg_cell = worksheet.cell(row=row, column=7) # Suggested
            
            real_original = str(df.iloc[row-2]['target_text']) if pd.notna(df.iloc[row-2]['target_text']) else ""
            suggestion = str(sugg_cell.value) if sugg_cell.value and str(sugg_cell.value) != 'None' else ""
            
            if suggestion and suggestion != "Mantido" and real_original != suggestion:
                matcher = difflib.SequenceMatcher(None, real_original, suggestion)
                rich_text_elements = []
                
                for tag, i1, i2, j1, j2 in matcher.get_opcodes():
                    text_part = suggestion[j1:j2]
                    if not text_part: continue
                    
                    if tag == 'equal':
                        rich_text_elements.append(text_part)
                    elif tag in ('insert', 'replace'):
                        rich_text_elements.append(TextBlock(font=red_font, text=text_part))
                
                has_changes = any(isinstance(x, TextBlock) for x in rich_text_elements)
                if has_changes:
                    sugg_cell.value = CellRichText(rich_text_elements)

    excel_data = base64.b64encode(output.getvalue()).decode()
    
    return {
        "excel_base64": excel_data,
        "filename": f"STMS_DB_Export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    }

@app.post("/stms/review_text")
def review_stms_text(item: STMSReviewRequest):
    # Endpoint genérico para scripts externos (como automacao_stms.py)
    mapping_item = {
        "string_name": "EXTERNAL_REQ",
        "en": item.source_text,
        "pt": item.target_text or "",
        "en_comment": item.context or "",
        "pt_comment": ""
    }
    if item.char_limit:
        mapping_item["en_comment"] += f" (MAX: {item.char_limit})"
        
    try:
        result = fetch_translation(mapping_item)
        advice = result.get('advice', item.target_text)
        if advice == "Mantido":
            advice = item.target_text
        return {"suggestion": advice, "reasoning": result.get('reason')}
    except Exception as e:
        return {"suggestion": item.target_text, "reasoning": str(e)}

# --- ROTAS DE TEAM BOARD ---

@app.get("/team-board/all")
def get_team_board_all(tab: str = "fixed", db: Session = Depends(get_db)):
    areas = db.query(TeamBoardArea).filter(TeamBoardArea.tab == tab).order_by(TeamBoardArea.position.asc()).all()
    area_ids = [a.id for a in areas]
    members = db.query(TeamBoardMember).filter(TeamBoardMember.area_id.in_(area_ids)).order_by(TeamBoardMember.position.asc()).all()
    
    user_ids = [m.user_id for m in members if m.user_id]
    users = db.query(User).filter(User.id.in_(user_ids)).all() if user_ids else []
    user_map = {u.id: u for u in users}

    members_data = []
    for m in members:
        m_dict = {
            "id": m.id, "area_id": m.area_id, "user_id": m.user_id,
            "name": m.name, "role": m.role, "identifier": m.identifier,
            "status": m.status, "prefix": m.prefix, "parent_id": m.parent_id,
            "is_highlighted": m.is_highlighted, "date_range": m.date_range,
            "position": m.position
        }
        if m.user_id and m.user_id in user_map:
            u = user_map[m.user_id]
            m_dict["name"] = f"{u.first_name} {u.last_name}".strip()
            m_dict["role"] = u.role if u.role else m.role
            m_dict["avatar"] = u.avatar
        members_data.append(m_dict)

    return {"areas": areas, "members": members_data}

class MemberPosition(BaseModel):
    id: int
    position: int

class TeamBoardBulkPositionUpdate(BaseModel):
    members: List[MemberPosition]

@app.post("/team-board-members/bulk-position")
def bulk_update_member_positions(data: TeamBoardBulkPositionUpdate, db: Session = Depends(get_db)):
    for item in data.members:
        db.query(TeamBoardMember).filter(TeamBoardMember.id == item.id).update({"position": item.position})
    db.commit()
    return {"status": "success"}

@app.post("/team-board/replicate-fixed")
def replicate_fixed_team(db: Session = Depends(get_db)):
    try:
        # 1. Delete all current areas, members, projects and their mappings
        # Cascade will handle most, but we explicitly clear projects for the current tab
        db.query(TeamBoardProject).filter(TeamBoardProject.tab == "current").delete()
        db.query(TeamBoardArea).filter(TeamBoardArea.tab == "current").delete()
        db.flush()
        
        # 2. Replicate Areas and Members from Fixed to Current
        fixed_areas = db.query(TeamBoardArea).filter(TeamBoardArea.tab == "fixed").all()
        member_id_map = {} # old_member_id -> new_member_id (from Fixed to Current)
        
        for area in fixed_areas:
            new_area = TeamBoardArea(
                name=area.name,
                section=area.section,
                position=area.position,
                tab="current"
            )
            db.add(new_area)
            db.flush()
            
            fixed_members = db.query(TeamBoardMember).filter(TeamBoardMember.area_id == area.id).order_by(TeamBoardMember.parent_id.asc().nullsfirst()).all()
            
            hierarchical_map = {} # fixed_id -> current_id (within this replication)
            
            for member in fixed_members:
                new_member = TeamBoardMember(
                    area_id=new_area.id,
                    name=member.name,
                    role=member.role,
                    identifier=member.identifier,
                    status=member.status,
                    prefix=member.prefix,
                    user_id=member.user_id,
                    parent_id=hierarchical_map.get(member.parent_id) if member.parent_id else None,
                    is_highlighted=member.is_highlighted,
                    date_range=member.date_range,
                    position=member.position
                )
                db.add(new_member)
                db.flush()
                hierarchical_map[member.id] = new_member.id
                member_id_map[member.id] = new_member.id
        
        # 3. Replicate Projects from Fixed to Current
        fixed_projects = db.query(TeamBoardProject).filter(TeamBoardProject.tab == "fixed").all()
        
        for proj in fixed_projects:
            new_proj = TeamBoardProject(
                name=proj.name,
                position=proj.position,
                tab="current"
            )
            db.add(new_proj)
            db.flush()
            
            # Replicate member mappings for this project
            fixed_mappings = db.query(TeamBoardProjectMember).filter(TeamBoardProjectMember.project_id == proj.id).all()
            for mapping in fixed_mappings:
                new_member_id = member_id_map.get(mapping.member_id)
                if new_member_id:
                    db.add(TeamBoardProjectMember(project_id=new_proj.id, member_id=new_member_id))
                
        db.commit()
        return {"status": "success"}
    except Exception as e:
        db.rollback()
        print(f"Erro ao replicar time: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/team-board-areas")
def create_team_board_area(area: TeamBoardAreaCreate, db: Session = Depends(get_db)):
    new_area = TeamBoardArea(**area.dict())
    db.add(new_area)
    db.commit()
    db.refresh(new_area)
    return new_area

@app.put("/team-board-areas/{area_id}")
def update_team_board_area(area_id: int, area: TeamBoardAreaCreate, db: Session = Depends(get_db)):
    db_area = db.query(TeamBoardArea).filter(TeamBoardArea.id == area_id).first()
    if not db_area:
        raise HTTPException(status_code=404, detail="Área não encontrada")
    for key, value in area.dict().items():
        setattr(db_area, key, value)
    db.commit()
    db.refresh(db_area)
    return db_area

class AreaPosition(BaseModel):
    id: int
    position: int

class TeamBoardAreaBulkPositionUpdate(BaseModel):
    areas: List[AreaPosition]

@app.post("/team-board-areas/bulk-position")
def bulk_update_area_positions(data: TeamBoardAreaBulkPositionUpdate, db: Session = Depends(get_db)):
    for item in data.areas:
        db.query(TeamBoardArea).filter(TeamBoardArea.id == item.id).update({"position": item.position})
    db.commit()
    return {"status": "success"}

@app.delete("/team-board-areas/{area_id}")
def delete_team_board_area(area_id: int, db: Session = Depends(get_db)):
    db_area = db.query(TeamBoardArea).filter(TeamBoardArea.id == area_id).first()
    if not db_area:
        raise HTTPException(status_code=404, detail="Área não encontrada")
    db.delete(db_area)
    db.commit()
    return {"status": "success"}

@app.post("/team-board-members")
def create_team_board_member(member: TeamBoardMemberCreate, db: Session = Depends(get_db)):
    new_member = TeamBoardMember(**member.dict())
    db.add(new_member)
    db.commit()
    db.refresh(new_member)
    return new_member

@app.put("/team-board-members/{member_id}")
def update_team_board_member(member_id: int, member: TeamBoardMemberCreate, db: Session = Depends(get_db)):
    db_member = db.query(TeamBoardMember).filter(TeamBoardMember.id == member_id).first()
    if not db_member:
        raise HTTPException(status_code=404, detail="Membro não encontrado")
        
    old_area_id = db_member.area_id
    
    for key, value in member.dict().items():
        setattr(db_member, key, value)
    db.commit()
    db.refresh(db_member)
    
    # Sincroniza a Célula do Perfil do usuário se a área for alterada no Time Semanal
    if db_member.user_id and db_member.area_id != old_area_id:
        new_area = db.query(TeamBoardArea).filter(TeamBoardArea.id == db_member.area_id).first()
        if new_area:
            db_user = db.query(User).filter(User.id == db_member.user_id).first()
            if db_user:
                db_user.cell = new_area.name
                db.commit()
                
    return db_member

@app.delete("/team-board-members/{member_id}")
def delete_team_board_member(member_id: int, db: Session = Depends(get_db)):
    db_member = db.query(TeamBoardMember).filter(TeamBoardMember.id == member_id).first()
    if not db_member:
        raise HTTPException(status_code=404, detail="Membro não encontrado")
    db.delete(db_member)
    db.commit()
    return {"status": "success"}

# --- ROTAS DE PROJETOS ---

class ProjectCreate(BaseModel):
    name: str
    tab: Optional[str] = "fixed"

class ProjectMemberUpdate(BaseModel):
    member_ids: List[int]

@app.get("/team-board/kp-projects")
def list_projects(tab: str = "fixed", db: Session = Depends(get_db)):
    projects = db.query(TeamBoardProject).filter(TeamBoardProject.tab == tab).order_by(TeamBoardProject.position.asc()).all()
    result = []
    for p in projects:
        # Get members through the mapping table
        member_ids = [m.member_id for m in db.query(TeamBoardProjectMember).filter(TeamBoardProjectMember.project_id == p.id).all()]
        result.append({
            "id": p.id,
            "name": p.name,
            "position": p.position,
            "tab": p.tab,
            "member_ids": member_ids,
            "created_at": p.created_at
        })
    return result

@app.post("/team-board/kp-projects")
def create_project(data: ProjectCreate, db: Session = Depends(get_db)):
    print(f"DEBUG: Criando projeto KP com nome: {data.name} na aba: {data.tab}")
    try:
        # Get last position in this tab
        last = db.query(TeamBoardProject).filter(TeamBoardProject.tab == data.tab).order_by(TeamBoardProject.position.desc()).first()
        next_pos = (last.position + 1) if last else 0
        
        new_project = TeamBoardProject(name=data.name, position=next_pos, tab=data.tab)
        db.add(new_project)
        db.commit()
        db.refresh(new_project)
        return new_project
    except Exception as e:
        db.rollback()
        print(f"ERRO ao criar projeto: {e}")
        raise HTTPException(status_code=500, detail=str(e))

class ProjectPosition(BaseModel):
    id: int
    position: int

class TeamBoardProjectBulkPositionUpdate(BaseModel):
    projects: List[ProjectPosition]

@app.post("/team-board/kp-projects/bulk-position")
def bulk_update_project_positions(data: TeamBoardProjectBulkPositionUpdate, db: Session = Depends(get_db)):
    for item in data.projects:
        db.query(TeamBoardProject).filter(TeamBoardProject.id == item.id).update({"position": item.position})
    db.commit()
    return {"status": "success"}

@app.put("/team-board/kp-projects/{project_id}/members")
def update_project_members(project_id: int, data: ProjectMemberUpdate, db: Session = Depends(get_db)):
    # Clear existing members
    db.query(TeamBoardProjectMember).filter(TeamBoardProjectMember.project_id == project_id).delete()
    # Add new members
    for mid in data.member_ids:
        db.add(TeamBoardProjectMember(project_id=project_id, member_id=mid))
    db.commit()
    return {"status": "success"}

@app.delete("/team-board/kp-projects/{project_id}")
def delete_project(project_id: int, db: Session = Depends(get_db)):
    db.query(TeamBoardProject).filter(TeamBoardProject.id == project_id).delete()
    db.commit()
    return {"status": "success"}

# --- ROTAS DE DAILY ISSUES ---

@app.get("/daily-projects")
def list_daily_projects(db: Session = Depends(get_db)):
    return db.query(DailyProject).order_by(DailyProject.name.asc()).all()

@app.post("/daily-projects")
def create_daily_project(data: DailyProjectCreate, db: Session = Depends(get_db)):
    existing = db.query(DailyProject).filter(DailyProject.name == data.name).first()
    if existing:
        raise HTTPException(status_code=400, detail="Já existe um projeto com este nome.")
    
    new_project = DailyProject(name=data.name)
    db.add(new_project)
    db.commit()
    db.refresh(new_project)
    return new_project

@app.delete("/daily-projects/{project_id}")
def delete_daily_project(project_id: int, db: Session = Depends(get_db)):
    project = db.query(DailyProject).filter(DailyProject.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Projeto não encontrado.")
    
    issues_count = db.query(DailyIssue).filter(DailyIssue.project_id == project_id).count()
    if issues_count > 0:
        raise HTTPException(status_code=400, detail="Não é possível excluir um projeto que possui issues registradas.")
    
    db.delete(project)
    db.commit()
    return {"message": "Projeto removido."}

# --- ROTAS DE AVISOS DO SISTEMA (SYSTEM NOTICES) ---

@app.get("/system-notices")
def get_system_notices(db: Session = Depends(get_db)):
    return db.query(SystemNotice).order_by(SystemNotice.created_at.desc()).all()

@app.get("/system-notices/active")
def get_active_system_notice(db: Session = Depends(get_db)):
    notice = db.query(SystemNotice).filter(SystemNotice.is_active == 1).order_by(SystemNotice.updated_at.desc()).first()
    return notice

@app.post("/system-notices")
def create_system_notice(data: SystemNoticeCreate, db: Session = Depends(get_db)):
    if data.is_active:
        db.query(SystemNotice).update({SystemNotice.is_active: 0})
        
    notice = SystemNotice(
        title=data.title,
        description=data.description,
        is_active=1 if data.is_active else 0
    )
    db.add(notice)
    db.commit()
    db.refresh(notice)
    return notice

@app.put("/system-notices/{notice_id}")
def update_system_notice(notice_id: int, data: SystemNoticeUpdate, db: Session = Depends(get_db)):
    notice = db.query(SystemNotice).filter(SystemNotice.id == notice_id).first()
    if not notice:
        raise HTTPException(status_code=404, detail="Aviso não encontrado")
        
    if data.is_active is not None:
        if data.is_active:
            db.query(SystemNotice).filter(SystemNotice.id != notice_id).update({SystemNotice.is_active: 0})
        notice.is_active = 1 if data.is_active else 0
        
    if data.title is not None:
        notice.title = data.title
    if data.description is not None:
        notice.description = data.description
        
    notice.updated_at = datetime.utcnow()
    
    db.commit()
    db.refresh(notice)
    return notice

@app.patch("/system-notices/{notice_id}/toggle")
def toggle_system_notice(notice_id: int, db: Session = Depends(get_db)):
    notice = db.query(SystemNotice).filter(SystemNotice.id == notice_id).first()
    if not notice:
        raise HTTPException(status_code=404, detail="Aviso não encontrado")
        
    if notice.is_active == 0:
        db.query(SystemNotice).update({SystemNotice.is_active: 0})
        notice.is_active = 1
        notice.updated_at = datetime.utcnow() # Update timestamp so clients show it again
    else:
        notice.is_active = 0
        
    db.commit()
    db.refresh(notice)
    return notice

@app.delete("/system-notices/{notice_id}")
def delete_system_notice(notice_id: int, db: Session = Depends(get_db)):
    notice = db.query(SystemNotice).filter(SystemNotice.id == notice_id).first()
    if not notice:
        raise HTTPException(status_code=404, detail="Aviso não encontrado")
        
    db.delete(notice)
    db.commit()
    return {"message": "Aviso deletado com sucesso"}

@app.get("/daily-issues")
def list_daily_issues(date: Optional[str] = None, db: Session = Depends(get_db)):
    query = db.query(DailyIssue)
    if date:
        query = query.filter(DailyIssue.date == date)
    return query.order_by(DailyIssue.created_at.desc()).all()

@app.post("/daily-issues")
def create_daily_issue(data: DailyIssueCreate, db: Session = Depends(get_db)):
    project = db.query(DailyProject).filter(DailyProject.id == data.projectId).first()
    if not project:
        raise HTTPException(status_code=404, detail="Projeto não encontrado.")
    
    new_issue = DailyIssue(
        project_id=data.projectId,
        title=data.title,
        date=data.date
    )
    db.add(new_issue)
    db.commit()
    db.refresh(new_issue)
    return new_issue

@app.post("/daily-issues/bulk")
def create_daily_issues_bulk(data: List[DailyIssueCreate], db: Session = Depends(get_db)):
    created_issues = []
    for item in data:
        project = db.query(DailyProject).filter(DailyProject.id == item.projectId).first()
        if not project:
            continue # Pula se o projeto não existir
            
        new_issue = DailyIssue(
            project_id=item.projectId,
            title=item.title,
            date=item.date
        )
        db.add(new_issue)
        created_issues.append(new_issue)
    
    db.commit()
    for issue in created_issues:
        db.refresh(issue)
    return created_issues

@app.delete("/daily-issues/{issue_id}")
def delete_daily_issue(issue_id: int, db: Session = Depends(get_db)):
    issue = db.query(DailyIssue).filter(DailyIssue.id == issue_id).first()
    if not issue:
        raise HTTPException(status_code=404, detail="Issue não encontrada.")
    
    db.delete(issue)
    db.commit()
    return {"message": "Issue removida."}

# --- ROTAS DE NOTIFICAÇÕES ---

@app.get("/notifications")
def list_notifications(db: Session = Depends(get_db), limit: int = 20):
    notifs = db.query(Notification).order_by(Notification.created_at.desc()).limit(limit).all()
    return notifs

@app.patch("/notifications/{notif_id}/read")
def mark_notification_as_read(notif_id: int, db: Session = Depends(get_db)):
    notif = db.query(Notification).filter(Notification.id == notif_id).first()
    if notif:
        notif.is_read = 1
        db.commit()
    return {"status": "ok"}

@app.delete("/notifications")
def clear_all_notifications(db: Session = Depends(get_db)):
    db.query(Notification).delete()
    db.commit()
    return {"status": "cleared"}

# --- ROTAS DE REMARK ISSUES ---

class IssueItem(BaseModel):
    type: str
    criticality: str
    issue_id: str
    description: str
    team: str
    app_name: Optional[str] = None

@app.post("/remark-issues")
def save_remark_issues(issues: List[IssueItem], db: Session = Depends(get_db)):
    for item in issues:
        new_issue = RemarkIssue(
            type=item.type,
            criticality=item.criticality,
            issue_id=item.issue_id,
            description=item.description,
            team=item.team,
            app_name=item.app_name
        )
        db.add(new_issue)
    db.commit()
    return {"status": "saved", "count": len(issues)}

@app.get("/remark-issues")
def get_remark_issues(team: Optional[str] = None, db: Session = Depends(get_db)):
    query = db.query(RemarkIssue)
    if team:
        query = query.filter(RemarkIssue.team == team)
    return query.order_by(RemarkIssue.created_at.desc()).all()

# --- PROXY PARA IA (SIDIA) ---

@app.post("/ai/analyze")
async def ai_analyze(payload: Dict[str, Any], db: Session = Depends(get_db)):
    messages = payload.get("messages", [])
    if not messages:
        return {"status": "error", "detail": "Nenhuma mensagem fornecida"}

    user_message = messages[-1].get("content", "").lower().strip()
    context = payload.get("context", {})
    tab = context.get("tab", "geral")
    
    is_ref_query = "referencia" in user_message or "referência" in user_message
    if is_ref_query:
        model_match = re.search(r"(sm-[a-z0-9_-]+|[a-z0-9_-]+tpa|[a-z0-9_-]+gto)", user_message)
        if model_match:
            raw_model = model_match.group(1).strip().upper()
            model_data = db.query(ReferenceModel).filter(or_(ReferenceModel.model_name.ilike(f"{raw_model}"), ReferenceModel.model_name.ilike(f"%{raw_model}%"))).first()
            if not model_data:
                clean_raw = re.sub(r'[^A-Z0-9]', '', raw_model)
                all_models = db.query(ReferenceModel).all()
                for m in all_models:
                    clean_db = re.sub(r'[^A-Z0-9]', '', m.model_name.upper())
                    if clean_raw == clean_db or clean_raw in clean_db or clean_db in clean_raw:
                        model_data = m
                        break
            if model_data:
                response_text = f"O modelo de referência para **{model_data.model_name}** é:\n\n**{model_data.ref_model}**"
                if tab == "modelo_referencia":
                    siblings = db.query(ReferenceModel).filter(ReferenceModel.ref_model == model_data.ref_model, ReferenceModel.model_name != model_data.model_name).limit(8).all()
                    if siblings: response_text += f"\n\n**Modelo Pai:** {model_data.ref_model}\n**Modelos Irmãos:** {', '.join([s.model_name for s in siblings])}"
                    else: response_text += f"\n\n**Modelo Pai:** {model_data.ref_model}"
                return {"message": {"role": "assistant", "content": response_text}}

    knowledge_dir = os.path.join(os.path.dirname(__file__), 'knowledge_base')
    knowledge_text = ""
    
    if os.path.exists(knowledge_dir):
        for filename in sorted(os.listdir(knowledge_dir)):
            if filename.endswith(".txt"):
                file_path = os.path.join(knowledge_dir, filename)
                try:
                    with open(file_path, 'r', encoding='utf-8') as f:
                        content = f.read()
                        knowledge_text += f"\n--- DOCUMENTO: {filename} ---\n{content}\n"
                except Exception as e:
                    print(f"Erro ao ler {filename}: {e}")

    chart_instruction = (
        "\n\nOPÇÃO DE VISUALIZAÇÃO: Você PODE e DEVE responder com gráficos quando houver dados numéricos, "
        "estatísticos ou comparativos. Para gerar um gráfico, utilize EXCLUSIVAMENTE o seguinte formato de bloco de código:\n"
        "```json:chart\n"
        "{\n"
        "  \"type\": \"bar\" | \"line\" | \"area\" | \"pie\",\n"
        "  \"title\": \"Título do Gráfico\",\n"
        "  \"data\": [\n"
        "    {\"name\": \"Categoria A\", \"value\": 10},\n"
        "    {\"name\": \"Categoria B\", \"value\": 20}\n"
        "  ]\n"
        "}\n"
        "```\n"
        "Use 'pie' para distribuições percentuais ou partes de um todo, 'bar' para comparações entre categorias, e 'line' para tendências temporais.\n"
    )

    ai_messages = [
        {
            "role": "system", 
            "content": (
                f"Você é o assistente inteligente do SVP (Specialized Verification Part). "
                f"Use a base de conhecimento abaixo para responder perguntas de forma natural.\n\n"
                f"REGRA CRÍTICA: NÃO use tabelas (Markdown tables) em suas respostas sob nenhuma circunstância. "
                f"Se precisar organizar dados, utilize listas com marcadores ou títulos em negrito."
                f"{chart_instruction}\n\n"
                f"Base de Conhecimento (NÃO CRIE TABELAS EM SUAS RESPOSTAS), CASO VOCÊ NÃO ACHE A INFORMAÇÃO. "
                f"DIGA QUE NÃO SABE, E NÃO INVENTE NADA:\n{knowledge_text}"
            )
        }
    ]
    
    for msg in messages:
        ai_messages.append({"role": msg.get("role", "user"), "content": msg.get("content", "")})

    ai_data = {
        "model": MODEL_ID,
        "messages": ai_messages,
        "stream": payload.get("stream", False)
    }

    try:
        response = requests.post(PROXY_URL, headers={"Authorization": f"Bearer {API_KEY}", "Content-Type": "application/json"}, json=ai_data, verify=False, timeout=120)
        if not response.ok: return {"status": "error", "source": "vLLM API", "code": response.status_code, "detail": response.text}
        return response.json()
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro interno no servidor: {str(e)}")

def parse_date(value: str, label: str) -> date:
    if not value or not value.strip():
        raise HTTPException(status_code=400, detail=f"{label} é obrigatória.")
    try:
        return datetime.strptime(value, "%Y-%m-%d").date()
    except ValueError:
        raise HTTPException(status_code=400, detail=f"{label} inválida. Use o formato AAAA-MM-DD.")

def periods_overlap(start_a: date, end_a: date, start_b: date, end_b: date) -> bool:
    return start_a <= end_b and end_a >= start_b

def authenticate_ldap(identifier: str, password: str) -> Optional[str]:
    if not identifier or not password:
        return None
    if Server is None or Connection is None:
        return None
    candidates = []
    if "@" in identifier:
        candidates.append(identifier)
    else:
        candidates.append(f"{identifier}@{LDAP_DOMAIN}")
        candidates.append(f"{LDAP_DOMAIN}\\{identifier}")

    for user in candidates:
        try:
            server = Server(LDAP_DC_HOST)
            conn = Connection(server, user=user, password=password, auto_bind=True)
            conn.unbind()
            return user # Retorna qual candidato funcionou
        except Exception:
            continue
    return None

# --- ROTAS ORIGINAIS ---

@app.post("/login")
def login(request: LoginRequest, db: Session = Depends(get_db)):
    identifier = (request.email or "").strip().lower()
    if Server is None or Connection is None:
        raise HTTPException(status_code=500, detail="Biblioteca LDAP não instalada no servidor.")
    
    auth_identifier = authenticate_ldap(identifier, request.password)
    if not auth_identifier:
        raise HTTPException(status_code=401, detail="Credenciais LDAP inválidas.")

    # Normalização robusta do identificador (ex: tira corp\ ou @corp)
    clean_id = identifier
    if "\\" in clean_id:
        clean_id = clean_id.split("\\")[-1]
    if "@" in clean_id:
        clean_id = clean_id.split("@")[0]
    
    # 1. Match por prefixo do email (ex: gilmar.silva@%) - MAIS ROBUSTO
    user = db.query(User).filter(User.email.ilike(f"{clean_id}@%")).first()
    
    # 2. Fallback match por nome completo aproximado (se prefixo falhar)
    if not user:
        clean_name_for_search = clean_id.replace('.', '').replace(' ', '').lower()
        all_users = db.query(User).all()
        for u in all_users:
            db_name_clean = f"{u.first_name}{u.last_name}".replace('.', '').replace(' ', '').lower()
            if db_name_clean == clean_name_for_search:
                user = u
                break

    if user:
        # Tenta descobrir a Célula pelo Time Semanal (apenas tab 'fixed')
        board_member = db.query(TeamBoardMember).join(TeamBoardArea).filter(
            TeamBoardMember.user_id == user.id,
            TeamBoardArea.tab == "fixed"
        ).first()
        
        cell_name = user.cell
        if board_member:
            area = db.query(TeamBoardArea).filter(TeamBoardArea.id == board_member.area_id).first()
            if area:
                cell_name = area.name

        return {
            "message": "Login realizado",
            "first_access": False,
            "user": {
                "id": user.id,
                "name": f"{user.first_name} {user.last_name}".strip(),
                "email": user.email,
                "role": user.role or "",
                "team": user.team or "",
                "cell": cell_name or "",
                "kp": user.kp if user.kp else (user.department or ""),
                "kp_type": user.kp_type,
                "sidia_id": user.department or "",
                "is_backup": bool(user.is_backup),
                "is_specialist": user.is_specialist,
                "avatar": user.avatar
            }
        }

    # SE NÃO EXISTE NO BANCO, CRIA AGORA (Auto-provisioning)
    # Usa o email que foi validado com sucesso no LDAP (normalizado para usuario@corp se necessário)
    final_email = auth_identifier.lower()
    if "\\" in final_email: 
        final_email = f"{final_email.split('\\')[-1]}@{LDAP_DOMAIN}"
    elif "@" not in final_email:
        final_email = f"{final_email}@{LDAP_DOMAIN}"
    
    display_name = clean_id.replace('.', ' ').title()
    name_parts = [p for p in re.split(r"[._\s]+", clean_id) if p]
    first_name = name_parts[0].capitalize() if name_parts else clean_id.capitalize()
    last_name = " ".join([p.capitalize() for p in name_parts[1:]]) if len(name_parts) > 1 else ""

    new_user = User(
        first_name=first_name,
        last_name=last_name,
        email=final_email,
        password="LDAP",
        role="LDAP",
        team="SVP",
        cell="SVP",
        kp="PENDENTE",
        department="000000",
        is_backup=0,
        is_specialist=0
    )

    try:
        db.add(new_user)
        db.commit()
        db.refresh(new_user)
        print(f"DEBUG: Novo usuário provisionado: {new_user.email} (ID: {new_user.id})")
    except Exception as e:
        db.rollback()
        # Se falhar a criação (ex: unique constraint que o match não pegou), 
        # tenta buscar uma última vez pelo email exato que causou o erro
        user = db.query(User).filter(User.email.ilike(final_email)).first()
        if user:
            return {
                "message": "Login realizado (Existente)",
                "first_access": False,
                "user": {
                    "id": user.id,
                    "name": f"{user.first_name} {user.last_name}".strip(),
                    "email": user.email,
                    "role": user.role or "",
                    "team": user.team or "",
                    "cell": user.cell or "",
                    "kp": user.kp or "",
                    "kp_type": user.kp_type,
                    "sidia_id": user.department or "",
                    "is_backup": bool(user.is_backup),
                    "is_specialist": user.is_specialist,
                    "avatar": user.avatar
                }
            }
        
        raise HTTPException(status_code=500, detail=f"Erro fatal ao criar usuário: {str(e)}")

    return {
        "message": "Login realizado (Novo usuário criado)",
        "first_access": True,
        "user": {
            "id": new_user.id,
            "name": f"{new_user.first_name} {new_user.last_name}".strip(),
            "email": new_user.email,
            "role": new_user.role or "",
            "team": new_user.team or "",
            "cell": new_user.cell or "",
            "kp": new_user.kp or "",
            "kp_type": new_user.kp_type,
            "sidia_id": new_user.department or "",
            "is_backup": bool(new_user.is_backup),
            "is_specialist": new_user.is_specialist,
            "avatar": new_user.avatar,
            "bio": new_user.bio,
            "skills": new_user.skills
        }
    }

@app.post("/users")
def create_user(payload: FirstAccessRequest, db: Session = Depends(get_db)):
    email = (payload.email or "").strip().lower()
    if not email:
        raise HTTPException(status_code=400, detail="E-mail é obrigatório.")

    if not payload.team or not payload.team.strip():
        raise HTTPException(status_code=400, detail="Time/Célula é obrigatório.")
    if not payload.kp or not payload.kp.strip():
        raise HTTPException(status_code=400, detail="KP de Projeto é obrigatório.")
    if not payload.role or not payload.role.strip():
        raise HTTPException(status_code=400, detail="Cargo/Título é obrigatório.")
    if not payload.sidia_id or not payload.sidia_id.strip():
        raise HTTPException(status_code=400, detail="Matrícula SIDIA é obrigatória.")

    # Busca insensível a maiúsculas/minúsculas
    user = db.query(User).filter(User.email.ilike(email)).first()
    
    base_name = (payload.display_name or email.split("@")[0]).strip()
    name_parts = [p for p in re.split(r"[._\s]+", base_name) if p]
    first_name = name_parts[0].capitalize() if name_parts else base_name.capitalize()
    last_name = " ".join([p.capitalize() for p in name_parts[1:]]) if len(name_parts) > 1 else ""

    if user:
        # UPSERT: Se já existe (pelo auto-provisioning do login), apenas atualiza os dados
        user.first_name = first_name
        user.last_name = last_name
        user.role = payload.role.strip()
        user.department = payload.sidia_id.strip()
        user.team = payload.team.strip()
        user.cell = (payload.cell or "").strip()
        user.kp = payload.kp.strip()
        user.kp_type = payload.kp_type
        user.is_backup = 1 if payload.is_backup else 0
        user.is_specialist = payload.is_specialist or 0
        if payload.avatar:
            user.avatar = payload.avatar
    else:
        # CREATE: Caso ainda não exista
        user = User(
            first_name=first_name,
            last_name=last_name,
            email=email.lower(),
            password="LDAP",
            role=payload.role.strip(),
            department=payload.sidia_id.strip(),
            team=payload.team.strip(),
            cell=(payload.cell or "").strip(),
            kp=payload.kp.strip(),
            kp_type=payload.kp_type,
            is_backup=1 if payload.is_backup else 0,
            is_specialist=payload.is_specialist or 0,
            avatar=payload.avatar
        )
        db.add(user)

    try:
        db.commit()
        db.refresh(user)
        
        # Sincroniza a área do Time Semanal se o usuário alterar a Célula no Perfil
        if user.cell:
            area = db.query(TeamBoardArea).filter(TeamBoardArea.name.ilike(user.cell)).first()
            if area:
                db.query(TeamBoardMember).filter(TeamBoardMember.user_id == user.id).update({"area_id": area.id})
                db.commit()
                
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Erro ao salvar usuário: {str(e)}")

    return {
        "message": "Usuário processado com sucesso.",
        "user": {
            "id": user.id,
            "name": f"{user.first_name} {user.last_name}".strip(),
            "email": user.email,
            "role": user.role or "",
            "team": user.team or "",
            "cell": user.cell or "",
            "kp": user.kp if user.kp else (user.department or ""),
            "kp_type": user.kp_type,
            "sidia_id": user.department or "",
            "is_backup": bool(user.is_backup),
            "is_specialist": user.is_specialist,
            "avatar": user.avatar,
            "bio": user.bio,
            "skills": user.skills
        }
    }

@app.get("/users")
def get_users(db: Session = Depends(get_db)):
    users = db.query(User).all()
    vacations = db.query(Vacation).all()
    
    # Buscar todas as áreas fixas dos membros
    board_members = db.query(TeamBoardMember, TeamBoardArea).join(TeamBoardArea).filter(TeamBoardArea.tab == "fixed").all()
    user_to_area = {m.TeamBoardMember.user_id: m.TeamBoardArea.name for m in board_members if m.TeamBoardMember.user_id}

    vacations_by_user: Dict[int, List[Dict[str, Any]]] = {}
    for v in vacations:
        if v.status == "rejected":
            continue
        vacations_by_user.setdefault(v.user_id, []).append({
            "start": v.start_date,
            "end": v.end_date,
            "status": v.status,
            "category": v.category
        })
    return [
        {
            "id": str(u.id),
            "name": f"{u.first_name} {u.last_name}",
            "role": u.role,
            "area": u.team,
            "cell": user_to_area.get(u.id, u.cell),
            "kp": u.kp if u.kp else u.department,
            "kp_type": u.kp_type,
            "is_backup": bool(u.is_backup),
            "email": u.email,
            "avatar": u.avatar,
            "bio": u.bio,
            "skills": u.skills,
            "vacations": vacations_by_user.get(u.id, [])
        }
        for u in users
    ]

@app.get("/users/{user_id}")
def get_user(user_id: int, db: Session = Depends(get_db)):
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="Usuário não encontrado")
    
    # Tenta descobrir a Célula pelo Time Semanal (apenas tab 'fixed') se aplicável
    board_member = db.query(TeamBoardMember).join(TeamBoardArea).filter(
        TeamBoardMember.user_id == user.id,
        TeamBoardArea.tab == "fixed"
    ).first()
    
    cell_name = user.cell
    if board_member:
        area = db.query(TeamBoardArea).filter(TeamBoardArea.id == board_member.area_id).first()
        if area:
            cell_name = area.name

    return {
        "id": user.id,
        "name": f"{user.first_name} {user.last_name}".strip(),
        "email": user.email,
        "role": user.role or "",
        "team": user.team or "",
        "cell": cell_name or "",
        "kp": user.kp if user.kp else (user.department or ""),
        "kp_type": user.kp_type,
        "sidia_id": user.department or "",
        "is_backup": bool(user.is_backup),
        "is_specialist": user.is_specialist,
        "avatar": user.avatar,
        "bio": user.bio or "",
        "skills": user.skills or ""
    }

@app.post("/vacations")
def request_vacations(request: VacationRequest, db: Session = Depends(get_db)):
    user = db.query(User).filter(User.id == request.userId).first()
    if not user:
        raise HTTPException(status_code=404, detail="Usuário não encontrado.")

    if not request.periods:
        raise HTTPException(status_code=400, detail="Informe ao menos um período de férias.")

    existing_vacations = db.query(Vacation).filter(
        Vacation.user_id == request.userId,
        Vacation.status.in_(["pending", "approved", "fluig_approved", "downloaded"])
    ).all()
    existing_periods = [
        (parse_date(v.start_date, "Data inicial"), parse_date(v.end_date, "Data final"))
        for v in existing_vacations
    ]

    user_team = (user.team or "").strip()
    
    group_periods: List[Tuple[date, date]] = []
    if user_team:
        other_team_members_ids = [
            row[0] for row in db.query(User.id)
            .filter(User.id != user.id)
            .filter(User.team == user_team)
            .all()
        ]
        
        if other_team_members_ids:
            group_vacations = db.query(Vacation).filter(
                Vacation.user_id.in_(other_team_members_ids),
                Vacation.status.in_(["pending", "approved", "fluig_approved", "downloaded"])
            ).all()
            group_periods = [
                (parse_date(v.start_date, "Data inicial"), parse_date(v.end_date, "Data final"))
                for v in group_vacations
            ]

    new_vacations: List[Vacation] = []
    new_periods = []
    for period in request.periods:
        start = parse_date(period.start, "Data inicial")
        end = parse_date(period.end, "Data final")
        if start > end:
            raise HTTPException(status_code=400, detail="Data final deve ser após a inicial.")

        for start_exist, end_exist in existing_periods:
            if periods_overlap(start, end, start_exist, end_exist):
                raise HTTPException(status_code=409, detail="Já existe férias pendente/aprovada nesse período.")

        for start_group, end_group in group_periods:
            if periods_overlap(start, end, start_group, end_group):
                raise HTTPException(
                    status_code=409,
                    detail="Membros da mesma célula/equipe não podem tirar férias no mesmo período."
                )

        for start_new, end_new in new_periods:
            if periods_overlap(start, end, start_new, end_new):
                raise HTTPException(status_code=400, detail="Os períodos informados se sobrepõem.")

        new_periods.append((start, end))
        vacation = Vacation(
            user_id=request.userId,
            start_date=start.isoformat(),
            end_date=end.isoformat(),
            category=period.category if period.category else "vacation",
            status=period.status if period.status else "pending",
            sell_days=1 if request.sellDays else 0
        )
        db.add(vacation)
        new_vacations.append(vacation)

    try:
        db.commit()
        for v in new_vacations:
            db.refresh(v)
        
        create_notification(
            db, 
            "Nova solicitação de férias", 
            f"{user.first_name} {user.last_name} solicitou férias para {len(new_vacations)} período(s).",
            "vacation",
            user_id=None
        )
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))

    return {
        "message": "Solicitação registrada.",
        "vacations": [
            {
                "id": v.id,
                "userId": v.user_id,
                "start": v.start_date,
                "end": v.end_date,
                "status": v.status
            }
            for v in new_vacations
        ]
    }

@app.get("/vacations")
def list_vacations(status: Optional[str] = None, user_id: Optional[int] = None, db: Session = Depends(get_db)):
    query = db.query(Vacation, User).join(User, Vacation.user_id == User.id)
    if status:
        statuses = [s.strip() for s in status.split(",") if s.strip()]
        if statuses:
            query = query.filter(Vacation.status.in_(statuses))
    if user_id:
        query = query.filter(Vacation.user_id == user_id)

    results = query.order_by(Vacation.start_date.asc()).all()
    return [
        {
            "id": v.id,
            "userId": v.user_id,
            "userName": f"{u.first_name} {u.last_name}",
            "userAvatar": u.avatar,
            "userRole": u.role,
            "team": u.team,
            "kp": u.kp if u.kp else u.department,
            "start": v.start_date,
            "end": v.end_date,
            "category": v.category,
            "status": v.status,
            "sellDays": bool(v.sell_days),
            "createdAt": v.created_at.isoformat() if v.created_at else None
        }
        for v, u in results
    ]

@app.delete("/vacations/{vacation_id}")
def delete_vacation(vacation_id: int, db: Session = Depends(get_db)):
    vacation = db.query(Vacation).filter(Vacation.id == vacation_id).first()
    if not vacation:
        raise HTTPException(status_code=404, detail="Férias não encontradas.")
    db.delete(vacation)
    db.commit()
    return {"message": "Registro de férias removido."}

@app.patch("/vacations/{vacation_id}")
def update_vacation_status(vacation_id: int, payload: VacationStatusUpdate, db: Session = Depends(get_db)):
    if payload.status not in {"pending", "approved", "rejected", "fluig_approved", "downloaded"}:
        raise HTTPException(status_code=400, detail="Status inválido.")
    
    vacation = db.query(Vacation).filter(Vacation.id == vacation_id).first()
    if not vacation:
        raise HTTPException(status_code=404, detail="Férias não encontradas.")
    
    if payload.status == "rejected":
        db.delete(vacation)
        db.commit()
        return {"id": vacation_id, "status": "deleted"}

    vacation.status = payload.status
    vacation.updated_at = datetime.utcnow()
    try:
        db.commit()
        db.refresh(vacation)
        user = db.query(User).filter(User.id == vacation.user_id).first()
        userName = f"{user.first_name} {user.last_name}" if user else "Usuário"
        status_pt = "aprovada" if payload.status == "approved" else "pendente"
        create_notification(db, f"Férias {status_pt}", f"A solicitação de férias de {userName} foi {status_pt}.", "vacation", user_id=vacation.user_id)
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    return {"id": vacation.id, "status": vacation.status}

@app.get("/reference-models")
def get_reference_models(
    page: int = 1, 
    limit: int = 15, 
    search: Optional[str] = None, 
    db: Session = Depends(get_db)
):
    query = db.query(ReferenceModel)
    if search:
        query = query.filter(
            or_(
                ReferenceModel.model_name.ilike(f"%{search}%"),
                ReferenceModel.ref_model.ilike(f"%{search}%")
            )
        )
    
    total = query.count()
    results = query.order_by(ReferenceModel.model_name.asc()).offset((page - 1) * limit).limit(limit).all()
    
    return {
        "total": total,
        "page": page,
        "limit": limit,
        "data": [
            {
                "id": r.id,
                "model_name": r.model_name,
                "ref_model": r.ref_model,
                "created_at": r.created_at.isoformat() if r.created_at else None
            }
            for r in results
        ]
    }

@app.post("/reference-models")
def create_reference_model(data: ReferenceModelCreate, db: Session = Depends(get_db)):
    # Check if model already exists
    exists = db.query(ReferenceModel).filter(ReferenceModel.model_name == data.model_name.strip().upper()).first()
    if exists:
        raise HTTPException(status_code=400, detail="Modelo já cadastrado.")
    
    new_model = ReferenceModel(
        model_name=data.model_name.strip().upper(),
        ref_model=data.ref_model.strip()
    )
    db.add(new_model)
    try:
        db.commit()
        db.refresh(new_model)
        return new_model
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))

@app.put("/reference-models/{model_id}")
def update_reference_model(model_id: int, data: ReferenceModelUpdate, db: Session = Depends(get_db)):
    model = db.query(ReferenceModel).filter(ReferenceModel.id == model_id).first()
    if not model:
        raise HTTPException(status_code=404, detail="Modelo não encontrado.")
    
    # Check if new model_name already exists for ANOTHER id
    exists = db.query(ReferenceModel).filter(ReferenceModel.model_name == data.model_name.strip().upper(), ReferenceModel.id != model_id).first()
    if exists:
        raise HTTPException(status_code=400, detail="Este nome de modelo já está em uso.")

    model.model_name = data.model_name.strip().upper()
    model.ref_model = data.ref_model.strip()
    
    try:
        db.commit()
        db.refresh(model)
        return model
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))

@app.delete("/reference-models/{model_id}")
def delete_reference_model(model_id: int, db: Session = Depends(get_db)):
    model = db.query(ReferenceModel).filter(ReferenceModel.id == model_id).first()
    if not model:
        raise HTTPException(status_code=404, detail="Modelo não encontrado.")
    
    db.delete(model)
    db.commit()
    return {"message": "Modelo removido com sucesso."}

def load_kb_data():
    kb_path = os.path.join(os.path.dirname(__file__), "knowledge_base", "remarks_data.json")
    if not os.path.exists(kb_path):
        return {'sampleId': [], 'account': [], 'samsungAccount': [], 'simCard': [], 'testerId': [], 'appName': [], 'deviceId': []}
    with open(kb_path, 'r', encoding='utf-8') as f:
        try:
            return json.load(f)
        except:
            return {'sampleId': [], 'account': [], 'samsungAccount': [], 'simCard': [], 'testerId': [], 'appName': [], 'deviceId': []}

def save_kb_data(data):
    kb_path = os.path.join(os.path.dirname(__file__), "knowledge_base", "remarks_data.json")
    with open(kb_path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

class KBEntry(BaseModel):
    field: str
    value: str

@app.get("/knowledge-base/remarks")
def get_kb_remarks(db: Session = Depends(get_db)):
    kb_data = load_kb_data()
    # Merge suggestions from the DB (Suggestion table) into the KB data
    try:
        all_suggestions = db.query(Suggestion).all()
        for s in all_suggestions:
            field = s.field_type
            value = s.value
            if field not in kb_data:
                kb_data[field] = []
            if not any(v.lower() == value.lower() for v in kb_data[field]):
                kb_data[field].append(value)
    except Exception as e:
        print(f"Aviso: Erro ao buscar sugestões do banco para KB remarks: {e}")
    return kb_data

@app.post("/knowledge-base/remarks")
def add_kb_remark(entry: KBEntry):
    data = load_kb_data()
    if entry.field not in data:
        data[entry.field] = []
    
    # Case-insensitive check
    if not any(val.lower() == entry.value.lower() for val in data[entry.field]):
        data[entry.field].append(entry.value)
        save_kb_data(data)
    return {"message": "Adicionado com sucesso", "data": data}

@app.delete("/knowledge-base/remarks")
def delete_kb_remark(field: str, value: str):
    data = load_kb_data()
    if field in data and value in data[field]:
        data[field].remove(value)
        save_kb_data(data)
    return {"message": "Removido com sucesso", "data": data}

@app.get("/search")
def search(field: str, query: str = "", db: Session = Depends(get_db)):
    if not query: return []
    
    results_list = []
    
    # 1. Busca no JSON KB
    kb_data = load_kb_data()
    if field in kb_data:
        for val in kb_data[field]:
            if query.lower() in val.lower() and val not in results_list:
                results_list.append(val)
                
    # Se for testerId, pesquisa também na tabela de usuários
    try:
        if field == "testerId":
            users = db.query(User).filter(
                or_(
                    User.first_name.ilike(f"%{query}%"),
                    User.last_name.ilike(f"%{query}%"),
                    User.email.ilike(f"%{query}%")
                )
            ).limit(10).all()
            for u in users:
                name = f"{u.first_name} {u.last_name}".strip()
                if name not in results_list:
                    results_list.append(name)
        
        db_results = db.query(Suggestion.value).filter(
            Suggestion.field_type == field, 
            Suggestion.value.ilike(f"%{query}%")
        ).limit(10).all()
        
        for r in db_results:
            if r[0] not in results_list:
                results_list.append(r[0])
    except Exception as e:
        print(f"Aviso: Erro ao buscar no banco de dados para a busca ({field}): {e}")
            
    return results_list[:10]

@app.post("/save")
def save_data(data: RemarkData, db: Session = Depends(get_db)):
    """
    1. Salva novas sugestões para autocomplete no banco e na knowledge base JSON.
    2. Salva o relatório completo no histórico.
    """
    
    # 1. Salvar Sugestões (Autocomplete) e no JSON
    fields_to_process = {
        'testerId': [data.testerId] + (data.testerIds or []),
        'account': [data.account] + (data.accounts or []),
        'samsungAccount': [data.samsungAccount] + (data.samsungAccounts or []),
        'simCard': [data.simCard] + (data.simCards or []),
        'appName': [data.appName] + (data.appNames or []),
        'sampleId': [data.sampleId] + (data.sampleIds or []),
        'deviceId': [data.deviceId] + (data.deviceIds or [])
    }

    kb_data = load_kb_data()
    kb_changed = False
    suggestions_saved = 0
    
    for field_name, values in fields_to_process.items():
        for value in values:
            if value and value.strip():
                clean_value = value.strip()
                # Salva no DB
                try:
                    exists = db.query(Suggestion).filter(
                        Suggestion.field_type == field_name,
                        Suggestion.value.ilike(clean_value)
                    ).first()
                    if not exists:
                        db.add(Suggestion(field_type=field_name, value=clean_value))
                        db.commit()
                        suggestions_saved += 1
                except Exception as e:
                    db.rollback()
                    print(f"Aviso: Não foi possível salvar a sugestão no banco de dados ({field_name}): {e}")
                
                # Salva na Knowledge Base (JSON)
                if field_name not in kb_data:
                    kb_data[field_name] = []
                
                # Case-insensitive check for JSON
                if not any(val.lower() == clean_value.lower() for val in kb_data[field_name]):
                    kb_data[field_name].append(clean_value)
                    kb_changed = True
                    
    if kb_changed:
        save_kb_data(kb_data)
    
    # 2. Salvar Relatório Completo (Apenas se tiver dados completos e testerId)
    if data.full_form_data and data.testerId:
        try:
            new_report = Report(
                team=data.team,
                tester_id=data.testerId,
                full_data=json.dumps(data.full_form_data)
            )
            db.add(new_report)
            db.commit()
        except Exception as e:
            print(f"Erro ao salvar histórico: {e}")
            db.rollback()
    
    return {"message": "Dados processados com sucesso.", "suggestions_added": suggestions_saved}

@app.post("/tickets")
def create_ticket(data: TicketCreate, db: Session = Depends(get_db)):
    new_ticket = Ticket(type=data.type, priority=data.priority, status=data.status if data.status else "pendente", title=data.title, content=data.content, resolution=data.resolution, creators=json.dumps(data.creators))
    db.add(new_ticket)
    db.commit()
    db.refresh(new_ticket)
    creator_names = ", ".join([c.get('name', 'Alguém') for c in data.creators])
    create_notification(db, f"Novo {data.type}: {data.title}", f"{creator_names} criou um novo {data.type}.", "ticket")
    return new_ticket

@app.get("/tickets")
def get_tickets(db: Session = Depends(get_db)):
    tickets = db.query(Ticket).order_by(Ticket.created_at.desc()).all()
    result = []
    for t in tickets:
        result.append({"id": t.id, "type": t.type, "priority": t.priority, "status": t.status, "title": t.title, "content": t.content, "resolution": t.resolution, "creators": json.loads(t.creators), "created_at": t.created_at.isoformat()})
    return result

@app.patch("/tickets/{ticket_id}/status")
def update_ticket_status(ticket_id: int, payload: TicketStatusUpdate, db: Session = Depends(get_db)):
    ticket = db.query(Ticket).filter(Ticket.id == ticket_id).first()
    if not ticket: raise HTTPException(status_code=404, detail="Ticket não encontrado")
    if payload.status: ticket.status = payload.status
    if payload.resolution: ticket.resolution = payload.resolution
    try:
        db.commit()
        db.refresh(ticket)
        create_notification(db, f"Status de {ticket.type} atualizado", f"O {ticket.type} '{ticket.title}' agora está como {ticket.status}.", "ticket")
        return {"id": ticket.id, "status": ticket.status}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))

# --- ROTAS DE GESTÃO DE CONHECIMENTO (KNOWLEDGE BASE) ---

class KnowledgeUpdateRequest(BaseModel):
    content: str

class SuggestKnowledgeRequest(BaseModel):
    content: str
    type: str

@app.get("/knowledge/{filename}")
def get_knowledge_file(filename: str):
    valid_files = ["bug_review.txt", "general_info.txt", "tone_of_voice.txt", "tone_of_voice_es.txt", "feedback_pt.txt", "feedback_es.txt"]
    if filename not in valid_files:
        raise HTTPException(status_code=400, detail="Arquivo inválido")
    
    file_path = os.path.join(os.path.dirname(__file__), "knowledge_base", filename)
    if not os.path.exists(file_path):
        return {"content": ""}
    
    with open(file_path, "r", encoding="utf-8") as f:
        return {"content": f.read()}

@app.post("/knowledge/{filename}")
def update_knowledge_file(filename: str, data: KnowledgeUpdateRequest):
    valid_files = ["bug_review.txt", "general_info.txt", "tone_of_voice.txt", "tone_of_voice_es.txt", "feedback_pt.txt", "feedback_es.txt"]
    if filename not in valid_files:
        raise HTTPException(status_code=400, detail="Arquivo inválido")
    
    kb_dir = os.path.join(os.path.dirname(__file__), "knowledge_base")
    os.makedirs(kb_dir, exist_ok=True)
    file_path = os.path.join(kb_dir, filename)
    
    try:
        with open(file_path, "w", encoding="utf-8") as f:
            f.write(data.content)
        return {"status": "success"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/ai/suggest_knowledge")
def suggest_knowledge(data: SuggestKnowledgeRequest):
    headers = {
        "Authorization": f"Bearer {API_KEY}",
        "Content-Type": "application/json"
    }
    
    prompt = f"O texto abaixo é um lote de itens da base de conhecimento ({data.type}) da Samsung.\n"
    prompt += "Sua tarefa é analisar criticamente esses itens e sugerir adições, correções ou melhorias técnicas.\n\n"
    prompt += "REGRAS DE FORMATAÇÃO (ESTRITAMENTE OBRIGATÓRIO):\n"
    prompt += "1. NÃO use tabelas Markdown. Use listas com marcadores (bullet points).\n"
    prompt += "2. Use títulos claros (## e ###) para organizar suas sugestões.\n"
    prompt += "3. Destaque termos técnicos importantes usando **negrito** ou `código`.\n"
    prompt += "4. Seja direto, profissional e focado em QA e Localização.\n\n"
    prompt += f"--- CONTEÚDO (LOTE) ---\n{data.content}"

    messages = [
        {"role": "system", "content": "Você é um consultor sênior de QA e Localização de Software. Suas respostas devem ser elegantes, organizadas em listas e nunca usar tabelas."},
        {"role": "user", "content": prompt}
    ]

    payload = {
        "model": MODEL_ID,
        "messages": messages,
        "stream": False
    }

    try:
        response = requests.post(PROXY_URL, json=payload, headers=headers, timeout=60, verify=False)
        if response.status_code == 200:
            res_data = response.json()
            suggestion = res_data.get('message', {}).get('content')
            if not suggestion and 'choices' in res_data and len(res_data['choices']) > 0:
                suggestion = res_data['choices'][0].get('message', {}).get('content', '')
            if not suggestion:
                suggestion = ''
            return {"suggestion": suggestion}
        else:
            raise HTTPException(status_code=500, detail="Erro na IA")
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# --- ROTAS DE GLOSSÁRIO ---

class GlossaryItem(BaseModel):
    id: Optional[str] = None
    english: str
    translation: str
    description: Optional[str] = ""
    dnt: Optional[str] = "No"
    app_name: Optional[str] = ""

@app.get("/glossary")
def list_glossary():
    conn = None
    try:
        conn = psycopg2.connect(**DB_CONFIG_XML)
        cur = conn.cursor(cursor_factory=DictCursor)
        cur.execute("SELECT id, english, translation, description, dnt, app_name FROM glossary ORDER BY english ASC")
        rows = cur.fetchall()
        return [dict(row) for row in rows]
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if conn: conn.close()

@app.post("/glossary")
def create_glossary_item(item: GlossaryItem):
    conn = None
    try:
        conn = psycopg2.connect(**DB_CONFIG_XML)
        cur = conn.cursor()
        item_id = item.id or str(int(time.time() * 1000))
        cur.execute("""
            INSERT INTO glossary (id, english, translation, description, dnt, app_name)
            VALUES (%s, %s, %s, %s, %s, %s)
        """, (item_id, item.english, item.translation, item.description, item.dnt, item.app_name))
        conn.commit()
        return {"status": "success", "id": item_id}
    except Exception as e:
        if conn: conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if conn: conn.close()

@app.put("/glossary/{item_id}")
def update_glossary_item(item_id: str, item: GlossaryItem):
    conn = None
    try:
        conn = psycopg2.connect(**DB_CONFIG_XML)
        cur = conn.cursor()
        cur.execute("""
            UPDATE glossary 
            SET english=%s, translation=%s, description=%s, dnt=%s, app_name=%s
            WHERE id=%s
        """, (item.english, item.translation, item.description, item.dnt, item.app_name, item_id))
        conn.commit()
        return {"status": "success"}
    except Exception as e:
        if conn: conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if conn: conn.close()

@app.delete("/glossary/{item_id}")
def delete_glossary_item(item_id: str):
    conn = None
    try:
        conn = psycopg2.connect(**DB_CONFIG_XML)
        cur = conn.cursor()
        cur.execute("DELETE FROM glossary WHERE id=%s", (item_id,))
        conn.commit()
        return {"status": "success"}
    except Exception as e:
        if conn: conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if conn: conn.close()

@app.get("/glossary_es")
def list_glossary_es():
    conn = None
    try:
        conn = psycopg2.connect(**DB_CONFIG_XML)
        cur = conn.cursor(cursor_factory=DictCursor)
        cur.execute("SELECT id, english, translation, description, dnt, app_name FROM glossary_es ORDER BY english ASC")
        rows = cur.fetchall()
        return [dict(row) for row in rows]
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if conn: conn.close()

@app.post("/glossary_es")
def create_glossary_item_es(item: GlossaryItem):
    conn = None
    try:
        conn = psycopg2.connect(**DB_CONFIG_XML)
        cur = conn.cursor()
        item_id = item.id or str(int(time.time() * 1000))
        cur.execute("""
            INSERT INTO glossary_es (id, english, translation, description, dnt, app_name)
            VALUES (%s, %s, %s, %s, %s, %s)
        """, (item_id, item.english, item.translation, item.description, item.dnt, item.app_name))
        conn.commit()
        return {"status": "success", "id": item_id}
    except Exception as e:
        if conn: conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if conn: conn.close()

@app.put("/glossary_es/{item_id}")
def update_glossary_item_es(item_id: str, item: GlossaryItem):
    conn = None
    try:
        conn = psycopg2.connect(**DB_CONFIG_XML)
        cur = conn.cursor()
        cur.execute("""
            UPDATE glossary_es 
            SET english=%s, translation=%s, description=%s, dnt=%s, app_name=%s
            WHERE id=%s
        """, (item.english, item.translation, item.description, item.dnt, item.app_name, item_id))
        conn.commit()
        return {"status": "success"}
    except Exception as e:
        if conn: conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if conn: conn.close()

@app.delete("/glossary_es/{item_id}")
def delete_glossary_item_es(item_id: str):
    conn = None
    try:
        conn = psycopg2.connect(**DB_CONFIG_XML)
        cur = conn.cursor()
        cur.execute("DELETE FROM glossary_es WHERE id=%s", (item_id,))
        conn.commit()
        return {"status": "success"}
    except Exception as e:
        if conn: conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if conn: conn.close()

class UserUpdateRequest(BaseModel):
    kp: Optional[str] = None
    kp_type: Optional[str] = None
    sidia_id: Optional[str] = None
    email: Optional[str] = None
    cell: Optional[str] = None
    name: Optional[str] = None
    role: Optional[str] = None
    team: Optional[str] = None
    is_backup: Optional[bool] = None
    is_specialist: Optional[bool] = None
    bio: Optional[str] = None
    skills: Optional[str] = None

@app.patch("/users/{user_id}")
def update_user_generic(user_id: int, req: UserUpdateRequest, db: Session = Depends(get_db)):
    try:
        user = db.query(User).filter(User.id == user_id).first()
        if not user:
            raise HTTPException(status_code=404, detail="Usuário não encontrado")
        
        # Valid fields to update
        req_data = req.dict(exclude_unset=True)
        allowed_fields = ['kp', 'kp_type', 'sidia_id', 'email', 'cell', 'name', 'role', 'team', 'is_backup', 'is_specialist', 'bio', 'skills']
        
        for key, value in req_data.items():
            if key in allowed_fields:
                if key == 'sidia_id':
                    user.department = value
                elif key == 'email':
                    # Check if email is already taken by someone else
                    existing = db.query(User).filter(User.email == value, User.id != user_id).first()
                    if existing:
                        raise HTTPException(status_code=400, detail="E-mail já está em uso")
                    user.email = value
                elif key == 'name':
                    import re
                    base_name = str(value).strip()
                    name_parts = [p for p in re.split(r"[._\s]+", base_name) if p]
                    user.first_name = name_parts[0].capitalize() if name_parts else base_name.capitalize()
                    user.last_name = " ".join([p.capitalize() for p in name_parts[1:]]) if len(name_parts) > 1 else ""
                elif key == 'is_backup':
                    user.is_backup = 1 if value else 0
                elif key == 'is_specialist':
                    user.is_specialist = 1 if value else 0
                else:
                    setattr(user, key, value)
                    
        db.commit()
        db.refresh(user)
        
        # Sincroniza a área do Time Semanal se o usuário alterar a Célula no Perfil via PATCH
        if 'cell' in req_data and user.cell:
            area = db.query(TeamBoardArea).filter(TeamBoardArea.name.ilike(user.cell)).first()
            if area:
                db.query(TeamBoardMember).filter(TeamBoardMember.user_id == user.id).update({"area_id": area.id})
                db.commit()
                
        return {"status": "success"}
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))

class UserUpdateEmailRequest(BaseModel):
    email: str

@app.patch("/users/{user_id}/email")
def update_user_email(user_id: int, req: UserUpdateEmailRequest, db: Session = Depends(get_db)):
    try:
        user = db.query(User).filter(User.id == user_id).first()
        if not user:
            raise HTTPException(status_code=404, detail="Usuário não encontrado")
        
        existing = db.query(User).filter(User.email == req.email, User.id != user_id).first()
        if existing:
            raise HTTPException(status_code=400, detail="E-mail já está em uso")
            
        user.email = req.email
        db.commit()
        return {"status": "success", "email": user.email}
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))

class UserUpdateAvatarRequest(BaseModel):
    avatar: str

@app.patch("/users/{user_id}/avatar")
def update_user_avatar(user_id: int, req: UserUpdateAvatarRequest, db: Session = Depends(get_db)):
    try:
        user = db.query(User).filter(User.id == user_id).first()
        if not user:
            raise HTTPException(status_code=404, detail="Usuário não encontrado")
        
        user.avatar = req.avatar
        db.commit()
        user.avatar = req.avatar
        db.commit()
        return {"status": "success", "avatar": user.avatar}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))

# --- SCHEMAS DE KANBAN ---
class KanbanCardCreate(BaseModel):
    title: str
    description: Optional[str] = None
    status: Optional[str] = "Backlog"
    type: str
    user_id: Optional[int] = None
    project_id: Optional[int] = None
    assigned_member_id: Optional[int] = None
    priority: Optional[str] = "Média"
    position: Optional[int] = 0
    deadline: Optional[str] = None

class KanbanCardUpdate(BaseModel):
    title: Optional[str] = None
    description: Optional[str] = None
    status: Optional[str] = None
    type: Optional[str] = None
    assigned_member_id: Optional[int] = None
    priority: Optional[str] = None
    position: Optional[int] = None
    deadline: Optional[str] = None

class KanbanCardPosition(BaseModel):
    id: int
    position: int
    status: Optional[str] = None

class KanbanCardBulkPositionUpdate(BaseModel):
    cards: List[KanbanCardPosition]

# --- ROTAS DE KANBAN ---
def find_board_member_helper(user, tab_type, db):
    # 1. Tentar por user_id
    member = db.query(TeamBoardMember).join(TeamBoardArea).filter(
        TeamBoardMember.user_id == user.id,
        TeamBoardArea.tab == tab_type
    ).first()
    if member:
        return member

    # 2. Tentar por nome completo
    full_name = f"{user.first_name} {user.last_name}".strip()
    member = db.query(TeamBoardMember).join(TeamBoardArea).filter(
        TeamBoardMember.name.ilike(full_name),
        TeamBoardArea.tab == tab_type
    ).first()
    if member:
        return member

    # 3. Carregar todos daquela aba para comparação Python mais robusta
    all_members = db.query(TeamBoardMember).join(TeamBoardArea).filter(
        TeamBoardArea.tab == tab_type
    ).all()
    
    u_first = user.first_name.lower().strip() if user.first_name else ""
    u_last = user.last_name.lower().strip() if user.last_name else ""
    u_email = user.email.lower().strip() if user.email else ""
    u_email_prefix = u_email.split('@')[0] if '@' in u_email else u_email
    
    # 3.1. Match por email prefix (identifier)
    if u_email_prefix:
        for m in all_members:
            if m.name and m.name.lower().strip() == u_email_prefix:
                return m
            if m.identifier and m.identifier.lower().strip() == u_email_prefix:
                return m
            if m.identifier and m.identifier.lower().strip() == u_email:
                return m
                
    # 3.2. Match contendo primeiro e último nome
    if u_first and u_last:
        for m in all_members:
            m_name_clean = m.name.lower() if m.name else ""
            if u_first in m_name_clean and u_last in m_name_clean:
                return m
                
    # 3.3. Match contendo primeiro nome como fallback
    if u_first:
        for m in all_members:
            m_name_clean = m.name.lower() if m.name else ""
            if u_first in m_name_clean:
                return m
                
    return None

@app.get("/kanban/squad")
def get_user_squad(user_id: int, db: Session = Depends(get_db)):
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="Usuário não encontrado.")
    
    # 1. Resolver o nome da Célula (prioridade para o perfil do usuário)
    cell_name = user.cell
    
    # Se o usuário não tiver célula no perfil, tentamos buscar a área em que ele está alocado no quadro semanal
    if not cell_name or not cell_name.strip():
        board_member = find_board_member_helper(user, "fixed", db)
            
        if board_member:
            area = db.query(TeamBoardArea).filter(TeamBoardArea.id == board_member.area_id).first()
            if area:
                cell_name = area.name

    # 2. Tentar encontrar projeto pelo nome da Célula (case-insensitive, apenas aba 'fixed')
    project = None
    if cell_name:
        cell_clean = cell_name.strip()
        project = db.query(TeamBoardProject).filter(
            TeamBoardProject.name.ilike(cell_clean),
            TeamBoardProject.tab == "fixed"
        ).first()
            
        # Se não encontrou a célula, criamos automaticamente
        if not project:
            try:
                last = db.query(TeamBoardProject).filter(TeamBoardProject.tab == "fixed").order_by(TeamBoardProject.position.desc()).first()
                next_pos = (last.position + 1) if last else 0
                project = TeamBoardProject(name=cell_clean, position=next_pos, tab="fixed")
                db.add(project)
                db.commit()
                db.refresh(project)
            except Exception as e:
                db.rollback()
                print(f"Aviso ao auto-criar projeto para célula: {e}")

    # 3. Se não encontrou por Célula, tenta encontrar projeto por KP
    if not project and user.kp:
        kp_clean = user.kp.strip()
        project = db.query(TeamBoardProject).filter(
            TeamBoardProject.name.ilike(kp_clean),
            TeamBoardProject.tab == "fixed"
        ).first()
            
        # Se não existe o projeto para o KP, criamos automaticamente na aba 'fixed'
        if not project:
            try:
                last = db.query(TeamBoardProject).filter(TeamBoardProject.tab == "fixed").order_by(TeamBoardProject.position.desc()).first()
                next_pos = (last.position + 1) if last else 0
                project = TeamBoardProject(name=kp_clean, position=next_pos, tab="fixed")
                db.add(project)
                db.commit()
                db.refresh(project)
            except Exception as e:
                db.rollback()
                print(f"Aviso ao auto-criar projeto para KP: {e}")
                
    if not project:
        raise HTTPException(
            status_code=404, 
            detail="Não conseguimos associar um time/squad para o seu usuário. Configure seu KP ou sua Célula em 'Meu Perfil'."
        )
        
    # 4. Obter membros do squad/célula
    members_data = []
    
    # Se o projeto é correspondente à Célula, preferimos pegar os membros alocados nela (TeamBoardArea)
    area = None
    if cell_name:
        area = db.query(TeamBoardArea).filter(
            TeamBoardArea.name.ilike(cell_name.strip()),
            TeamBoardArea.tab == "fixed"
        ).first()
            
    if area:
        # Se for um time baseado em Célula, pegamos os membros alocados nela no quadro
        cell_members = db.query(TeamBoardMember).filter(TeamBoardMember.area_id == area.id).all()
        user_ids = [m.user_id for m in cell_members if m.user_id]
        users = db.query(User).filter(User.id.in_(user_ids)).all() if user_ids else []
        user_map = {u.id: u for u in users}
        
        all_users = db.query(User).all()
        
        for m in cell_members:
            avatar = None
            resolved_uid = m.user_id
            
            # Tentar resolver user_id se for nulo
            if not resolved_uid and m.name:
                m_name_clean = m.name.lower().strip()
                m_ident_clean = m.identifier.lower().strip() if m.identifier else ""
                
                for u in all_users:
                    u_first = u.first_name.lower().strip() if u.first_name else ""
                    u_last = u.last_name.lower().strip() if u.last_name else ""
                    u_full = f"{u_first} {u_last}"
                    u_email = u.email.lower().strip() if u.email else ""
                    u_email_prefix = u_email.split('@')[0] if '@' in u_email else u_email
                    
                    if m_ident_clean and (m_ident_clean == u_email or m_ident_clean == u_email_prefix):
                        resolved_uid = u.id
                        avatar = u.avatar
                        break
                    elif m_name_clean == u_full:
                        resolved_uid = u.id
                        avatar = u.avatar
                        break
                    elif u_first and u_last and u_first in m_name_clean and u_last in m_name_clean:
                        resolved_uid = u.id
                        avatar = u.avatar
                        break
            
            if resolved_uid and not avatar and resolved_uid in user_map:
                avatar = user_map[resolved_uid].avatar
                
            members_data.append({
                "id": m.id,
                "name": m.name,
                "role": m.role,
                "user_id": resolved_uid,
                "avatar": avatar
            })
    else:
        # Caso contrário (projeto de KP puro), pegamos os membros do projeto associados no Time Semanal
        proj_members = db.query(TeamBoardProjectMember).filter(TeamBoardProjectMember.project_id == project.id).all()
        member_ids = [pm.member_id for pm in proj_members]
        
        members = db.query(TeamBoardMember).filter(TeamBoardMember.id.in_(member_ids)).all() if member_ids else []
        user_ids = [m.user_id for m in members if m.user_id]
        users = db.query(User).filter(User.id.in_(user_ids)).all() if user_ids else []
        user_map = {u.id: u for u in users}
        
        all_users = db.query(User).all()
        
        for m in members:
            avatar = None
            resolved_uid = m.user_id
            
            if not resolved_uid and m.name:
                m_name_clean = m.name.lower().strip()
                m_ident_clean = m.identifier.lower().strip() if m.identifier else ""
                
                for u in all_users:
                    u_first = u.first_name.lower().strip() if u.first_name else ""
                    u_last = u.last_name.lower().strip() if u.last_name else ""
                    u_full = f"{u_first} {u_last}"
                    u_email = u.email.lower().strip() if u.email else ""
                    u_email_prefix = u_email.split('@')[0] if '@' in u_email else u_email
                    
                    if m_ident_clean and (m_ident_clean == u_email or m_ident_clean == u_email_prefix):
                        resolved_uid = u.id
                        avatar = u.avatar
                        break
                    elif m_name_clean == u_full:
                        resolved_uid = u.id
                        avatar = u.avatar
                        break
                    elif u_first and u_last and u_first in m_name_clean and u_last in m_name_clean:
                        resolved_uid = u.id
                        avatar = u.avatar
                        break
                        
            if resolved_uid and not avatar and resolved_uid in user_map:
                avatar = user_map[resolved_uid].avatar
                
            members_data.append({
                "id": m.id,
                "name": m.name,
                "role": m.role,
                "user_id": resolved_uid,
                "avatar": avatar
            })
            
    # Garantir que o próprio usuário solicitante esteja na lista de membros para fins de delegação
    user_member_ids = [m["user_id"] for m in members_data if m["user_id"]]
    if user.id not in user_member_ids:
        self_member = find_board_member_helper(user, "fixed", db)
            
        if self_member:
            members_data.append({
                "id": self_member.id,
                "name": self_member.name,
                "role": self_member.role,
                "user_id": user.id,
                "avatar": user.avatar
            })
        else:
            members_data.append({
                "id": 999000 + user.id,
                "name": f"{user.first_name} {user.last_name}".strip(),
                "role": user.role,
                "user_id": user.id,
                "avatar": user.avatar
            })
            
    return {
        "id": project.id,
        "name": project.name,
        "members": members_data
    }

@app.get("/kanban/cards")
def get_kanban_cards(type: str, user_id: Optional[int] = None, project_id: Optional[int] = None, db: Session = Depends(get_db)):
    if type == "pessoal":
        if not user_id:
            raise HTTPException(status_code=400, detail="user_id é obrigatório para Kanban pessoal")
        return db.query(KanbanCard).filter(KanbanCard.type == "pessoal", KanbanCard.user_id == user_id).order_by(KanbanCard.position.asc(), KanbanCard.created_at.desc()).all()
    elif type == "time":
        if project_id:
            project = db.query(TeamBoardProject).filter(TeamBoardProject.id == project_id).first()
            if project:
                same_name_projects = db.query(TeamBoardProject.id).filter(TeamBoardProject.name.ilike(project.name)).all()
                project_ids = [p[0] for p in same_name_projects]
                return db.query(KanbanCard).filter(
                    KanbanCard.type == "time", 
                    KanbanCard.project_id.in_(project_ids)
                ).order_by(KanbanCard.position.asc(), KanbanCard.created_at.desc()).all()
            return db.query(KanbanCard).filter(KanbanCard.type == "time", KanbanCard.project_id == project_id).order_by(KanbanCard.position.asc(), KanbanCard.created_at.desc()).all()
        
        if not user_id:
            raise HTTPException(status_code=400, detail="user_id ou project_id é obrigatório para Kanban de time")
        
        user = db.query(User).filter(User.id == user_id).first()
        if not user:
            return []
        
        # 1. Tentar resolver célula (prioridade para o perfil)
        cell_name = user.cell
        
        # Se não tiver célula no perfil, busca no time semanal
        if not cell_name or not cell_name.strip():
            board_member = find_board_member_helper(user, "fixed", db)
                
            if board_member:
                area = db.query(TeamBoardArea).filter(TeamBoardArea.id == board_member.area_id).first()
                if area:
                    cell_name = area.name

        # 2. Tentar obter pelo KP (case-insensitive, apenas aba 'fixed')
        project = None
        if user.kp:
            kp_clean = user.kp.strip()
            project = db.query(TeamBoardProject).filter(
                TeamBoardProject.name.ilike(kp_clean),
                TeamBoardProject.tab == "fixed"
            ).first()
            
        # 3. Tentar obter pela célula (case-insensitive, apenas aba 'fixed')
        if not project and cell_name:
            cell_clean = cell_name.strip()
            project = db.query(TeamBoardProject).filter(
                TeamBoardProject.name.ilike(cell_clean),
                TeamBoardProject.tab == "fixed"
            ).first()
            
        if not project:
            return []
            
        # --- RECOVERY DE CARDS ÓRFÃOS ---
        # Se existem cards do time criados por este usuário que perderam o project_id
        # (devido a replicações passadas), vamos reconectá-los ao projeto fixo atual.
        orphan_cards = db.query(KanbanCard).filter(
            KanbanCard.type == "time",
            KanbanCard.project_id == None,
            KanbanCard.user_id == user.id
        ).all()
        if orphan_cards:
            for card in orphan_cards:
                card.project_id = project.id
            db.commit()
        # --------------------------------
            
        same_name_projects = db.query(TeamBoardProject.id).filter(TeamBoardProject.name.ilike(project.name)).all()
        project_ids = [p[0] for p in same_name_projects]
        return db.query(KanbanCard).filter(
            KanbanCard.type == "time", 
            KanbanCard.project_id.in_(project_ids)
        ).order_by(KanbanCard.position.asc(), KanbanCard.created_at.desc()).all()
    else:
        raise HTTPException(status_code=400, detail="Tipo inválido. Deve ser 'pessoal' ou 'time'")

@app.post("/kanban/cards")
def create_kanban_card(card: KanbanCardCreate, db: Session = Depends(get_db)):
    db_card = KanbanCard(
        title=card.title,
        description=card.description,
        status=card.status or "Backlog",
        type=card.type,
        user_id=card.user_id,
        project_id=card.project_id,
        assigned_member_id=card.assigned_member_id,
        priority=card.priority or "Média",
        position=card.position or 0,
        deadline=card.deadline
    )
    db.add(db_card)
    db.commit()
    db.refresh(db_card)
    return db_card

@app.post("/kanban/cards/bulk-position")
def bulk_update_kanban_card_positions(data: KanbanCardBulkPositionUpdate, db: Session = Depends(get_db)):
    try:
        for item in data.cards:
            update_data = {"position": item.position}
            if item.status is not None:
                update_data["status"] = item.status
            db.query(KanbanCard).filter(KanbanCard.id == item.id).update(update_data)
        db.commit()
        return {"status": "success"}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))

@app.put("/kanban/cards/{card_id}")
def update_kanban_card(card_id: int, card: KanbanCardUpdate, db: Session = Depends(get_db)):
    db_card = db.query(KanbanCard).filter(KanbanCard.id == card_id).first()
    if not db_card:
        raise HTTPException(status_code=404, detail="Card não encontrado")
    
    if card.title is not None:
        db_card.title = card.title
    if card.description is not None:
        db_card.description = card.description
    if card.status is not None:
        db_card.status = card.status
    if card.type is not None:
        db_card.type = card.type
    if card.assigned_member_id is not None:
        db_card.assigned_member_id = None if card.assigned_member_id == 0 else card.assigned_member_id
    if card.priority is not None:
        db_card.priority = card.priority
    if card.position is not None:
        db_card.position = card.position
    if card.deadline is not None:
        db_card.deadline = card.deadline
        
    db.commit()
    db.refresh(db_card)
    return db_card

@app.delete("/kanban/cards/{card_id}")
def delete_kanban_card(card_id: int, db: Session = Depends(get_db)):
    db_card = db.query(KanbanCard).filter(KanbanCard.id == card_id).first()
    if not db_card:
        raise HTTPException(status_code=404, detail="Card não encontrado")
    db.delete(db_card)
    db.commit()
    return {"status": "success", "message": "Card removido com sucesso"}

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8001)